from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import fitz
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


SHEET_6183 = "6183"
SHEET_TEREKE = "tereke"
READING_6183 = "6183 Yazıları Okuma"
READING_TEREKE = "Tereke Yazıları Okuma"

OUTPUT_COLUMNS = [
    "gonderen_kurum",
    "sayi",
    "tarih",
    "kisi_ad_soyad_unvan",
    "tckn_vkn",
    "haciz_konulan_tutar",
    "hesap_no",
    "hbno",
    "pdf_dosyasi",
]

BALANCE_INFO_COLUMN = "HESAP TÜRÜ"
ACCOUNT_CHECK_COLUMN = "hesap_kontrol"
INTERNAL_COLUMNS = [
    *OUTPUT_COLUMNS,
    "kisi_turu",
    BALANCE_INFO_COLUMN,
    ACCOUNT_CHECK_COLUMN,
    "bakiye_bilgi",
]

SHEET_6183_COLUMNS = [
    None,
    "kisi_ad_soyad_unvan",
    None,
    "kisi_turu",
    "tckn_vkn",
    "gonderen_kurum",
    "tarih",
    "sayi",
    None,
    "haciz_konulan_tutar",
    "hbno",
    "hesap_no",
    BALANCE_INFO_COLUMN,
    None,
    None,
    None,
    None,
    None,
    None,
    None,
    None,
    None,
    "pdf_dosyasi",
    ACCOUNT_CHECK_COLUMN,
]


@dataclass
class EypRecord:
    gonderen_kurum: str
    sayi: str
    tarih: str
    kisi_ad_soyad_unvan: str
    tckn_vkn: str
    haciz_konulan_tutar: str
    hesap_no: str
    hesap_turu: str
    hbno: str
    pdf_dosyasi: str
    kisi_turu: str = ""
    HESAP_TURU: str = ""
    hesap_kontrol: str = ""

    def as_dict(self) -> dict[str, str]:
        values = {col: getattr(self, col) for col in OUTPUT_COLUMNS}
        values["kisi_turu"] = self.kisi_turu
        values[BALANCE_INFO_COLUMN] = self.HESAP_TURU
        values[ACCOUNT_CHECK_COLUMN] = self.hesap_kontrol
        values["bakiye_bilgi"] = self.HESAP_TURU
        return values


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def label_key(value: str) -> str:
    value = normalize_space(value).casefold()
    value = "".join(
        char for char in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(char)
    )
    return value


def format_sender_for_sheet(sheet_title: str, value) -> str:
    sender = normalize_space(str(value or ""))
    if not sender:
        return ""
    if sheet_title != SHEET_6183:
        return sender

    return re.sub(
        r"\b(defter(?:dar|tar)l[ıi]ğ[ıi])\b\s*",
        r"\1\n",
        sender,
        count=1,
        flags=re.IGNORECASE,
    )


def pdf_text(pdf_path: Path) -> str:
    with fitz.open(pdf_path) as doc:
        return "\n".join(page.get_text() for page in doc)


def pdf_table_records(pdf_path: Path, defaults: dict[str, str]) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    with fitz.open(pdf_path) as doc:
        for page in doc:
            words = page.get_text("words")
            header = find_table_header(words)
            if not header:
                header = find_flexible_hbno_table_header(words)
            if header:
                records.extend(extract_records_from_words(words, header, defaults))
                continue

            transfer_header = find_transfer_table_header(words)
            if transfer_header:
                records.extend(extract_transfer_records_from_words(words, transfer_header, defaults))
                continue

            taxpayer_header = find_taxpayer_grid_header(words)
            if taxpayer_header:
                records.extend(extract_taxpayer_grid_records_from_words(words, taxpayer_header, defaults))
    return records


def find_table_header(words: list[tuple]) -> dict[str, float] | None:
    hbno_words = [word for word in words if word[4] == "HBNO"]
    for hbno_word in hbno_words:
        _, header_y, _, _, _, *_ = hbno_word
        header_words = [
            word for word in words if abs(word[1] - header_y) < 12 or abs(word[3] - header_y) < 16
        ]
        positions = {"hbno": hbno_word[0]}
        sorted_header_words = sorted(header_words, key=lambda item: item[0])
        for index, word in enumerate(sorted_header_words):
            x0, _, _, _, text, *_ = word
            if text == "VKN":
                positions["vkn"] = x0
            elif text == "TCKN":
                positions["tckn"] = x0
            elif text.startswith("AD-SOYAD"):
                positions["name"] = x0
            elif text.startswith("ŞUBE"):
                positions["branch"] = x0
            elif text == "BANKA" and index + 1 < len(sorted_header_words):
                next_text = sorted_header_words[index + 1][4]
                if next_text == "ADI":
                    positions["bank"] = x0
            elif text == "HESAP" and index + 1 < len(sorted_header_words):
                next_text = sorted_header_words[index + 1][4]
                if next_text == "NO":
                    positions["account"] = x0
                elif next_text == "TÜRÜ":
                    positions["account_type"] = x0
            elif text == "BLOKE":
                positions["amount"] = x0

        required = {"hbno", "vkn", "tckn", "name", "account", "amount"}
        if required.issubset(positions):
            positions["header_y"] = header_y
            return positions
    return None


def find_flexible_hbno_table_header(words: list[tuple]) -> dict[str, float] | None:
    hbno_words = [word for word in words if word[4] == "HBNO"]
    for hbno_word in hbno_words:
        _, header_y, _, _, _, *_ = hbno_word
        header_words = [
            word for word in words if abs(word[1] - header_y) < 18 or abs(word[3] - header_y) < 22
        ]
        positions = {"hbno": hbno_word[0], "header_y": header_y}
        sorted_header_words = sorted(header_words, key=lambda item: item[0])
        for index, word in enumerate(sorted_header_words):
            x0, _, _, _, text, *_ = word
            if text == "VKN":
                positions["vkn"] = x0
            elif text == "TCKN":
                positions["tckn"] = x0
            elif text.startswith("AD-SOYAD") or text.startswith("SOYAD/UNVA") or text == "AD-":
                positions["name"] = x0
            elif text.startswith("ŞUBE"):
                positions["branch"] = x0
            elif text == "BANKA" and index + 1 < len(sorted_header_words):
                next_text = sorted_header_words[index + 1][4]
                if next_text == "ADI":
                    positions["bank"] = x0
            elif text == "HESAP":
                positions["account"] = x0
            elif text == "TUTAR" or text == "BLOKE":
                positions["amount"] = x0

        required = {"hbno", "vkn", "tckn", "name", "account", "amount"}
        if required.issubset(positions):
            return positions
    return None


def find_transfer_table_header(words: list[tuple]) -> dict[str, float] | None:
    identity_words = [
        word for word in words if word[4].startswith("T.C/VERGİ") or word[4] == "MÜKELLEF"
    ]
    for identity_word in identity_words:
        _, header_y, _, _, _, *_ = identity_word
        header_words = [
            word for word in words if abs(word[1] - header_y) < 12 or abs(word[3] - header_y) < 16
        ]
        positions = {"identity": identity_word[0]}
        sorted_header_words = sorted(header_words, key=lambda item: item[0])
        for index, word in enumerate(sorted_header_words):
            x0, _, _, _, text, *_ = word
            if text.startswith("SOYAD-AD"):
                positions["name"] = x0
            elif text.startswith("AD-SOYAD"):
                positions["name"] = x0
            elif text.startswith("H.BİLDİRİ"):
                positions["hbno"] = x0
            elif text == "HACİZ":
                positions["hbno"] = x0
            elif text.startswith("ŞUBE"):
                positions["branch"] = x0
            elif text == "BANKA" and index + 1 < len(sorted_header_words):
                nearby_texts = [word[4] for word in sorted_header_words[index + 1 : index + 4]]
                if "HESAP" in nearby_texts:
                    positions["account"] = x0
                elif "ŞUBE" in nearby_texts:
                    positions["branch"] = x0
            elif text.startswith("AKTARILACAK"):
                positions["amount"] = x0

        required = {"identity", "name", "hbno", "account", "branch", "amount"}
        if required.issubset(positions):
            positions["header_y"] = header_y
            return positions
    return None


def find_taxpayer_grid_header(words: list[tuple]) -> dict[str, float] | None:
    marker_words = [word for word in words if word[4] == "MÜKELLEFİN"]
    for marker in marker_words:
        _, marker_y, _, _, _, *_ = marker
        header_words = [word for word in words if marker_y < word[1] < marker_y + 45]
        if not header_words:
            continue

        positions = {"header_y": max(word[1] for word in header_words)}
        for index, word in enumerate(sorted(header_words, key=lambda item: item[0])):
            x0, _, _, _, text, *_ = word
            next_words = [item[4] for item in sorted(header_words, key=lambda item: item[0])[index + 1 : index + 4]]
            if text == "BANKA" and "ADI" in next_words:
                positions["bank"] = x0
            elif text == "ADI" and "SOYADI" in next_words:
                positions["name"] = x0
            elif text == "HESAP":
                positions["account"] = x0
            elif text == "AKTARILACAK":
                positions["amount"] = x0
            elif text == "HACİZ":
                positions["hbno"] = x0
            elif text == "VERGİ":
                positions["vkn"] = x0
            elif text == "TC":
                positions["tckn"] = x0

        required = {"name", "account", "amount", "hbno", "vkn", "tckn"}
        if required.issubset(positions):
            return positions
    return None


def extract_records_from_words(
    words: list[tuple], header: dict[str, float], defaults: dict[str, str]
) -> list[dict[str, str]]:
    rows: dict[float, list[tuple]] = {}
    stop_y = find_signature_start_y(words, header["header_y"])
    for word in words:
        x0, y0, _, _, text, *_ = word
        if y0 <= header["header_y"] + 8 or y0 >= stop_y:
            continue
        first_column_x = min(header["hbno"], header["vkn"], header["tckn"], header["name"])
        if x0 < first_column_x - 5 or text in {"Vakıf", "ALICI"}:
            continue
        row_y = find_existing_row_y(rows, y0)
        rows.setdefault(row_y, []).append(word)

    records: list[dict[str, str]] = []
    current: dict[str, str] | None = None
    current_y = 0.0
    for row_y in sorted(rows):
        columns = table_row_columns(rows[row_y], header)
        if not any(columns.values()):
            continue

        has_identifier = bool(
            re.fullmatch(r"\d{10}", columns["vkn"])
            or re.fullmatch(r"\d{11}", columns["tckn"])
            or re.search(r"\d", columns["hbno"])
        )
        if has_identifier:
            if current:
                records.append(current)
            current = {
                "kisi_ad_soyad_unvan": columns["name"],
                "tckn": columns["tckn"] or defaults.get("keywords_tckn", ""),
                "vkn": columns["vkn"] or defaults.get("keywords_vkn", ""),
                "haciz_konulan_tutar": columns["amount"],
                "hesap_no": columns["account"],
                "hesap_turu": columns["account_type"] or "TL",
                "hbno": columns["hbno"],
                "banka_adi": "",
                "sube_adi": columns["branch"],
            }
            current_y = row_y
        elif current and row_y - current_y <= 12:
            append_table_continuation(current, columns)

    if current:
        records.append(current)
    return records


def extract_taxpayer_grid_records_from_words(
    words: list[tuple], header: dict[str, float], defaults: dict[str, str]
) -> list[dict[str, str]]:
    rows: dict[float, list[tuple]] = {}
    stop_y = min(
        find_signature_start_y(words, header["header_y"]),
        find_after_table_stop_y(words, header["header_y"]),
    )
    for word in words:
        x0, y0, _, _, text, *_ = word
        if y0 <= header["header_y"] + 8 or y0 >= stop_y:
            continue
        first_column_x = min(value for key, value in header.items() if key != "header_y")
        if x0 < first_column_x - 5:
            continue
        row_y = find_existing_row_y(rows, y0)
        rows.setdefault(row_y, []).append(word)

    records = []
    current: dict[str, str] | None = None
    for row_y in sorted(rows):
        columns = taxpayer_grid_row_columns(rows[row_y], header)
        if not any(columns.values()):
            continue

        has_record = bool(
            columns["account"]
            and columns["amount"]
            and columns["hbno"]
            and (columns["vkn"] or columns["tckn"])
        )
        if has_record:
            if current:
                records.append(current)
            identity_tckn = normalize_identity(columns["tckn"])
            identity_vkn = normalize_identity(columns["vkn"])
            current = {
                "kisi_ad_soyad_unvan": columns["name"],
                "tckn": identity_tckn or defaults.get("keywords_tckn", ""),
                "vkn": identity_vkn or defaults.get("keywords_vkn", ""),
                "haciz_konulan_tutar": columns["amount"],
                "hesap_no": columns["account"],
                "hesap_turu": "TL",
                "hbno": columns["hbno"],
                "banka_adi": columns["bank"],
                "sube_adi": "",
            }
        elif current and columns["name"]:
            current["kisi_ad_soyad_unvan"] = normalize_space(
                f"{current['kisi_ad_soyad_unvan']} {columns['name']}"
            )

    if current:
        records.append(current)
    return records


def extract_bank_transfer_grid_records(
    pdf_path: Path,
    defaults: dict[str, str],
) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    with fitz.open(pdf_path) as doc:
        for page in doc:
            words = [word for word in page.get_text("words") if normalize_space(word[4])]
            header = find_bank_transfer_grid_header(words)
            if not header:
                continue
            records.extend(extract_bank_transfer_grid_page_records(words, header, defaults))
    return dedupe_record_dicts(records)


def find_bank_transfer_grid_header(words: list[tuple]) -> dict[str, float] | None:
    for bank_word in words:
        if label_key(bank_word[4]) != "banka":
            continue
        _, header_y, _, _, _, *_ = bank_word
        header_words = [
            word
            for word in words
            if abs(word[1] - header_y) < 12 or abs(word[3] - header_y) < 16
        ]
        sorted_header_words = sorted(header_words, key=lambda item: item[0])
        positions = {"bank": bank_word[0], "header_y": header_y}
        for index, word in enumerate(sorted_header_words):
            x0, _, _, _, text, *_ = word
            key = label_key(text).replace(".", "")
            next_keys = [
                label_key(item[4]).replace(".", "")
                for item in sorted_header_words[index + 1 : index + 4]
            ]
            if key == "soyadi":
                positions["surname"] = x0
            elif key == "adi" and "banka" not in next_keys:
                positions["name"] = x0
            elif key == "sube":
                positions["branch"] = x0
            elif key == "iban":
                positions["account"] = x0
            elif key == "miktar":
                positions["amount"] = x0
            elif key == "hb":
                positions["hbno"] = x0
            elif key == "nevi":
                positions["levy_type"] = x0
            elif key == "vergi":
                positions["vkn"] = x0
            elif key.startswith("tckimlik") or (key == "tc" and "no" in next_keys):
                positions["tckn"] = x0

        required = {"bank", "surname", "name", "branch", "account", "amount", "hbno", "vkn"}
        if required.issubset(positions):
            positions.setdefault("tckn", positions["vkn"] + 40)
            return positions
    return None


def extract_bank_transfer_grid_page_records(
    words: list[tuple],
    header: dict[str, float],
    defaults: dict[str, str],
) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    identity_words = [
        word
        for word in words
        if word[1] > header["header_y"] + 20
        and word[0] >= header["vkn"] - 10
        and re.fullmatch(r"\d{10,11}", normalize_identity(word[4]))
    ]
    identity_words.sort(key=lambda item: item[1])

    for index, identity_word in enumerate(identity_words):
        row_y = identity_word[1]
        top = row_y - 18
        bottom = identity_words[index + 1][1] - 18 if index + 1 < len(identity_words) else row_y + 24
        stop_y = find_signature_start_y(words, header["header_y"])
        bottom = min(bottom, stop_y)
        row_words = [
            word
            for word in words
            if top <= word[1] <= bottom and word[1] > header["header_y"] + 20
        ]
        record = bank_transfer_grid_row_record(row_words, header, defaults)
        if record and (record["tckn"] or record["vkn"]):
            records.append(record)

    return records


def bank_transfer_grid_row_record(
    words: list[tuple],
    header: dict[str, float],
    defaults: dict[str, str],
) -> dict[str, str] | None:
    columns = bank_transfer_grid_row_columns(words, header)
    tckn = normalize_identity(columns["tckn"])
    vkn = normalize_identity(columns["vkn"])
    if not tckn and not vkn:
        return None

    return {
        "kisi_ad_soyad_unvan": normalize_space(f"{columns['surname']} {columns['name']}"),
        "tckn": tckn if len(tckn) == 11 else defaults.get("keywords_tckn", ""),
        "vkn": vkn if len(vkn) == 10 else defaults.get("keywords_vkn", ""),
        "haciz_konulan_tutar": columns["amount"],
        "hesap_no": columns["account"],
        "hesap_turu": "TL",
        "hbno": columns["hbno"],
        "banka_adi": columns["bank"],
        "sube_adi": columns["branch"],
    }


def bank_transfer_grid_row_columns(words: list[tuple], header: dict[str, float]) -> dict[str, str]:
    column_positions = {
        "bank": header["bank"],
        "surname": header["surname"],
        "name": header["name"],
        "branch": header["branch"],
        "account": header["account"],
        "amount": header["amount"],
        "hbno": header["hbno"],
        "levy_type": header.get("levy_type", header["vkn"] - 25),
        "vkn": header["vkn"],
        "tckn": header["tckn"],
    }
    sorted_columns = sorted(column_positions.items(), key=lambda item: item[1])
    ranges = {}
    for index, (column, x0) in enumerate(sorted_columns):
        start = (
            (sorted_columns[index - 1][1] + x0) / 2
            if index > 0
            else x0 - 20
        )
        end = (
            (x0 + sorted_columns[index + 1][1]) / 2
            if index + 1 < len(sorted_columns)
            else 10_000
        )
        ranges[column] = (start, end)

    columns = {name: [] for name in column_positions}
    for word in sorted(words, key=lambda item: item[0]):
        x0, _, _, _, _, *_ = word
        for column, (start, end) in ranges.items():
            if start <= x0 < end:
                columns[column].append(word)
                break

    return {
        name: normalize_space(" ".join(word[4] for word in reading_order(values)))
        for name, values in columns.items()
    }


def find_after_table_stop_y(words: list[tuple], header_y: float) -> float:
    stop_markers = {"VERGİ", "5(*)T.C.vatandaşı"}
    candidates = [
        word[1]
        for word in words
        if word[1] > header_y and (word[4] in stop_markers or word[4].startswith("5(*)"))
    ]
    return min(candidates) if candidates else 10_000


def extract_transfer_records_from_words(
    words: list[tuple], header: dict[str, float], defaults: dict[str, str]
) -> list[dict[str, str]]:
    rows: dict[float, list[tuple]] = {}
    stop_y = find_signature_start_y(words, header["header_y"])
    for word in words:
        x0, y0, _, _, _, *_ = word
        if y0 <= header["header_y"] + 16 or y0 >= stop_y:
            continue
        if x0 < header["identity"] - 5:
            continue
        row_y = find_existing_row_y(rows, y0)
        rows.setdefault(row_y, []).append(word)

    records: list[dict[str, str]] = []
    current: dict[str, str] | None = None
    current_y = 0.0
    pending_account_parts: list[str] = []
    for row_y in sorted(rows):
        columns = transfer_table_row_columns(rows[row_y], header)
        identity = columns["identity"]
        if not re.fullmatch(r"\d{10,11}", identity):
            if columns["account"]:
                if current and row_y - current_y <= 15:
                    current["hesap_no"] = normalize_space(
                        f"{current['hesap_no']} {columns['account']}"
                    )
                elif not current:
                    pending_account_parts.append(columns["account"])
            continue

        if current:
            records.append(current)
        account_parts = [*pending_account_parts]
        if columns["account"]:
            account_parts.append(columns["account"])
        pending_account_parts = []
        current = {
            "kisi_ad_soyad_unvan": columns["name"],
            "tckn": identity if len(identity) == 11 else defaults.get("keywords_tckn", ""),
            "vkn": identity if len(identity) == 10 else defaults.get("keywords_vkn", ""),
            "haciz_konulan_tutar": columns["amount"],
            "hesap_no": normalize_space(" ".join(account_parts)),
            "hesap_turu": "TL",
            "hbno": columns["hbno"],
            "banka_adi": "",
            "sube_adi": columns["branch"],
        }
        current_y = row_y

    if current:
        records.append(current)
    return records


def find_signature_start_y(words: list[tuple], header_y: float) -> float:
    candidates = []
    for word in words:
        _, y0, _, _, text, *_ = word
        if y0 <= header_y:
            continue
        if text == "Bu" or text in {"Vergi", "Müdürü", "Müdür", "Yardımcısı"}:
            candidates.append(y0)
    return min(candidates) if candidates else 10_000


def find_existing_row_y(rows: dict[float, list[tuple]], y0: float) -> float:
    for row_y in rows:
        if abs(row_y - y0) <= 8:
            return row_y
    return y0


def table_row_columns(words: list[tuple], header: dict[str, float]) -> dict[str, str]:
    column_positions = {
        "hbno": header["hbno"],
        "vkn": header["vkn"],
        "tckn": header["tckn"],
        "name": header["name"],
        "account": header["account"],
        "amount": header["amount"],
    }
    if "branch" in header:
        column_positions["branch"] = header["branch"]
    if "bank" in header:
        column_positions["bank"] = header["bank"]
    if "account_type" in header:
        column_positions["account_type"] = header["account_type"]

    sorted_columns = sorted(column_positions.items(), key=lambda item: item[1])
    ranges = {}
    for index, (column, x0) in enumerate(sorted_columns):
        end = sorted_columns[index + 1][1] - 5 if index + 1 < len(sorted_columns) else 10_000
        ranges[column] = (x0 - 5, end)

    columns = {
        "hbno": [],
        "vkn": [],
        "tckn": [],
        "name": [],
        "bank": [],
        "branch": [],
        "account": [],
        "account_type": [],
        "amount": [],
    }
    for word in sorted(words, key=lambda item: item[0]):
        x0, _, _, _, text, *_ = word
        for column, (start, end) in ranges.items():
            if start <= x0 < end:
                columns[column].append(word)
                break
    return {name: normalize_space(" ".join(word[4] for word in reading_order(values))) for name, values in columns.items()}


def reading_order(words: list[tuple]) -> list[tuple]:
    return sorted(words, key=lambda word: (round(word[1] / 8), word[0]))


def transfer_table_row_columns(words: list[tuple], header: dict[str, float]) -> dict[str, str]:
    if header["branch"] < header["account"]:
        branch_range = (header["branch"] - 5, header["hbno"] - 5)
        account_range = (header["account"] - 5, header["amount"] - 5)
    else:
        account_range = (header["account"] - 5, header["branch"] - 5)
        branch_range = (header["branch"] - 5, header["amount"] - 5)

    ranges = {
        "identity": (header["identity"] - 5, header["name"] - 5),
        "name": (header["name"] - 5, header["hbno"] - 5),
        "hbno": (header["hbno"] - 5, header["account"] - 5),
        "account": account_range,
        "branch": branch_range,
        "amount": (header["amount"] - 5, 10_000),
    }
    columns = {name: [] for name in ranges}
    for word in sorted(words, key=lambda item: item[0]):
        x0, _, _, _, text, *_ = word
        for column, (start, end) in ranges.items():
            if start <= x0 < end:
                columns[column].append(text)
                break
    return {name: normalize_space(" ".join(values)) for name, values in columns.items()}


def taxpayer_grid_row_columns(words: list[tuple], header: dict[str, float]) -> dict[str, str]:
    column_positions = {
        "bank": header.get("bank", 0),
        "name": header["name"],
        "account": header["account"],
        "amount": header["amount"],
        "hbno": header["hbno"],
        "vkn": header["vkn"],
        "tckn": header["tckn"],
    }
    sorted_columns = sorted(column_positions.items(), key=lambda item: item[1])
    ranges = {}
    for index, (column, x0) in enumerate(sorted_columns):
        end = sorted_columns[index + 1][1] - 5 if index + 1 < len(sorted_columns) else 10_000
        ranges[column] = (x0 - 5, end)

    columns = {name: [] for name in column_positions}
    for word in sorted(words, key=lambda item: item[0]):
        x0, _, _, _, text, *_ = word
        for column, (start, end) in ranges.items():
            if start <= x0 < end:
                columns[column].append(word)
                break
    return {
        name: normalize_space(" ".join(word[4] for word in reading_order(values)))
        for name, values in columns.items()
    }


def append_table_continuation(record: dict[str, str], columns: dict[str, str]) -> None:
    if columns["name"]:
        record["kisi_ad_soyad_unvan"] = normalize_space(
            f"{record['kisi_ad_soyad_unvan']} {columns['name']}"
        )
    if columns["account"]:
        record["hesap_no"] = normalize_space(
            f"{record['hesap_no']} {columns['account']}"
        )
    if columns["branch"]:
        record["sube_adi"] = normalize_space(
            f"{record['sube_adi']} {columns['branch']}"
        )
    if columns["account_type"]:
        record["hesap_turu"] = normalize_space(
            f"{record['hesap_turu']} {columns['account_type']}"
        )
    if columns["amount"] and not record["haciz_konulan_tutar"]:
        record["haciz_konulan_tutar"] = columns["amount"]


def identity_number_for_record(record: dict[str, str]) -> str:
    if is_legal_entity(record.get("kisi_ad_soyad_unvan", "")):
        return record.get("vkn", "") or record.get("tckn", "")
    return record.get("tckn", "") or record.get("vkn", "")


def person_type_for_record(record: dict[str, str]) -> str:
    if is_legal_entity(record.get("kisi_ad_soyad_unvan", "")):
        return "Tüzel"
    return "Gerçek"


def format_name_or_title(name: str) -> str:
    name = normalize_space(name)
    if not name:
        return ""
    if is_legal_entity(name):
        return format_legal_entity_name(name)

    words = [word for word in name.split() if not re.search(r"\d", word)]
    if not words:
        return ""
    if len(words) == 1:
        return upper_tr(words[0])

    first_names = " ".join(title_tr(word) for word in words[:-1])
    surname = upper_tr(words[-1])
    return normalize_space(f"{first_names} {surname}")


def format_legal_entity_name(value: str) -> str:
    normalized = normalize_space(value)
    normalized = re.sub(r"\s*\.\s*", ".", normalized)
    normalized = re.sub(r"\s{2,}", " ", normalized)

    lowercase_words = {"ve", "veya", "ile", "için", "adına", "adına", "da", "de"}
    abbreviation_words = {"A.Ş.", "A.Ş", "A.S.", "A.S", "LTD", "LTD.", "STI", "ŞTİ"}

    parts = re.findall(r"[A-Za-zÇĞİÖŞÜçğıöşü0-9\.]+|[^A-Za-zÇĞİÖŞÜçğıöşü0-9\.]+", normalized)
    formatted_parts: list[str] = []

    for part in parts:
        if re.fullmatch(r"[^A-Za-zÇĞİÖŞÜçğıöşü0-9\.]+", part):
            formatted_parts.append(part)
            continue

        if part.upper() in abbreviation_words:
            formatted_parts.append(part.upper())
            continue

        clean_part = part.strip(".")
        if clean_part.lower() in lowercase_words:
            formatted_parts.append(lower_tr(part))
            continue

        formatted_parts.append(title_tr(part))

    return normalize_space("".join(formatted_parts))


def lower_tr(value: str) -> str:
    return value.translate(str.maketrans("IİĞÜŞÖÇ", "ıiğüşöç")).lower()


def upper_tr(value: str) -> str:
    return value.upper()


def title_tr(value: str) -> str:
    lowered = lower_tr(value)
    return upper_tr(lowered[:1]) + lowered[1:]


def is_legal_entity(name: str) -> bool:
    normalized = normalize_space(name).upper()
    normalized_label = label_key(name)
    if re.search(
        r"\b(a\.?s|anonim|limited|ltd|sti|sirket|ticaret|tic|sanayi|san|insaat|ins|gida|turizm)\b",
        normalized_label,
    ):
        return True
    legal_markers = (
        " A.Ş",
        " A.S",
        " ANONİM",
        " LIMITED",
        " LTD",
        " ŞTİ",
        " STI",
        " KOOPERATİF",
        " DERNEĞİ",
        " VAKFI",
        " BELEDİYESİ",
        " BAŞKANLIĞI",
        " MÜDÜRLÜĞÜ",
        " TİCARET",
        " SANAYİ",
        " İNŞAAT",
        " TURİZM",
        " GIDA",
        " TEKSTİL",
        " OTOMOTİV",
        " ŞİRKET",
        " ŞİRKETİ",
    )
    return any(marker in f" {normalized} " for marker in legal_markers)


def get_pdf_header_fields(text: str) -> dict[str, str]:
    fields = {"sayi": "", "tarih": "", "gonderen_kurum": ""}
    lines = [normalize_space(line) for line in text.splitlines() if normalize_space(line)]

    for index, line in enumerate(lines):
        if line.lower() == "sayı" and index + 1 < len(lines):
            fields["sayi"] = lines[index + 1].lstrip(":").strip()
        elif re.match(r"^:?E-\d", line):
            fields["sayi"] = line.lstrip(":").strip()

        date_match = re.search(r"\b\d{2}\.\d{2}\.\d{4}\b", line)
        if date_match and not fields["tarih"]:
            fields["tarih"] = date_match.group(0)

    if "T.C." in lines:
        start = lines.index("T.C.")
        sender_lines = []
        for line in lines[start + 1 : start + 6]:
            if line.lower() in {"sayı", "konu"} or re.match(r"^:?E-", line):
                break
            if line in {"T.C.", "GELİR İDARESİ BAŞKANLIĞI"}:
                continue
            if not re.search(r"\d{2}\.\d{2}\.\d{4}", line):
                sender_lines.append(line)
        fields["gonderen_kurum"] = " ".join(sender_lines)

    if fields["sayi"]:
        fields["sayi"] = extract_document_number(fields["sayi"])
        # Tarih pattern'ini sayıdan çıkar
        date_match = re.search(r"\b\d{2}\.\d{2}\.\d{4}\b", fields["sayi"])
        if date_match:
            fields["sayi"] = fields["sayi"].replace(date_match.group(0), "").strip()
            if not fields["tarih"]:
                fields["tarih"] = date_match.group(0)

    return fields


def extract_document_number(value: str) -> str:
    value = normalize_space(value).lstrip(":").strip()
    if "-" not in value:
        return value
    return value.rsplit("-", 1)[-1].strip()


def extract_amounts(text: str) -> list[str]:
    amounts = []
    for match in re.finditer(r"\b\d{1,3}(?:\.\d{3})*,\d{2}\b", text):
        value = match.group(0)
        if value not in amounts:
            amounts.append(value)
    return amounts


def extract_vertical_taxpayer_records(text: str, defaults: dict[str, str]) -> list[dict[str, str]]:
    lines = [normalize_space(line) for line in text.splitlines() if normalize_space(line)]
    start = find_line_index(lines, "Mükellefin")
    if start is None:
        return []

    section = lines[start:]
    identity_value = taxpayer_identity_value(section)
    vkn, tckn = extract_vkn_tckn_from_text(identity_value)

    name = value_after_label(section, "Adı Soyadı (Ünvanı)")
    account_no = value_after_label(section, "Hesap Numarası")
    hbno = value_after_label_sequence(section, ["Haciz", "Bildirisinin", "Sayısı"])
    if not hbno:
        hbno = value_after_label(section, "Sayısı")
    amount = value_after_label_sequence(section, ["Aktarılacak", "Tutar"])
    if not amount:
        amount = value_after_label(section, "Aktarılacak Tutar")

    if not any((vkn, tckn, name, account_no, hbno, amount)):
        return []

    return [
        {
            "kisi_ad_soyad_unvan": name,
            "tckn": tckn or defaults.get("keywords_tckn", ""),
            "vkn": vkn or defaults.get("keywords_vkn", ""),
            "haciz_konulan_tutar": amount,
            "hesap_no": account_no,
            "hesap_turu": "TL",
            "hbno": hbno,
            "banka_adi": "",
            "sube_adi": "",
        }
    ]


def taxpayer_identity_value(lines: list[str]) -> str:
    for label in (
        "TC/Vergi Kimlik Numarası",
        "Vergi Kimlik Numarası/T.C.",
        "Vergi Kimlik Numarası / T.C.",
        "T.C. / Vergi Kimlik Numarası",
    ):
        value = value_after_label(lines, label)
        if value:
            return value

    for index, line in enumerate(lines):
        normalized = label_key(line)
        if "vergi kimlik numarasi" in normalized and ("tc" in normalized or "t.c" in normalized):
            return lines[index + 1] if index + 1 < len(lines) else ""

    return ""


def extract_horizontal_records(text: str, defaults: dict[str, str]) -> list[dict[str, str]]:
    lines = [normalize_space(line) for line in text.splitlines() if normalize_space(line)]
    section_start = find_line_index(lines, "MÜKELLEFİN")
    if section_start is None:
        return []

    section_end = find_line_index(lines, "VERGİ DAİRESİNİN", start=section_start + 1)
    section = lines[section_start + 1 : section_end] if section_end is not None else lines[section_start + 1 :]

    identity_line = find_label_line(section, "Vergi Kimlik Numarası") or find_label_line(section, "T.C. Kimlik Numarası")
    vkn, tckn = extract_vkn_tckn_from_text(identity_line) if identity_line else ("", "")
    name = value_after_label(section, "Adı Soyadı / Unvanı")
    identity = choose_identity_for_entity(name, vkn, tckn)
    if not identity:
        identity = value_after_label(section, "Vergi Kimlik Numarası") or value_after_label(section, "T.C. Kimlik Numarası")

    account_no = value_after_label(section, "Banka Hesap Numarası")
    amount = value_after_label(section, "Aktarılacak Tutar")
    hbno = value_after_label(section, "Haciz Bildirisi Numarası")

    if not any((identity, name, account_no, amount, hbno)):
        return []

    return [
        {
            "kisi_ad_soyad_unvan": name,
            "tckn": identity if len(normalize_identity(identity)) == 11 else defaults.get("keywords_tckn", ""),
            "vkn": identity if len(normalize_identity(identity)) == 10 else defaults.get("keywords_vkn", ""),
            "haciz_konulan_tutar": amount,
            "hesap_no": account_no,
            "hesap_turu": "TL",
            "hbno": hbno,
            "banka_adi": "",
            "sube_adi": "",
        }
    ]


def extract_taxpayer_text_grid_records(text: str, defaults: dict[str, str]) -> list[dict[str, str]]:
    lines = [normalize_space(line) for line in text.splitlines() if normalize_space(line)]
    try:
        start = lines.index("MÜKELLEFİN")
    except ValueError:
        return []

    try:
        stop = lines.index("VERGİ DAİRESİNİN", start + 1)
    except ValueError:
        stop = len(lines)
    section = lines[start + 1 : stop]

    if not {"BANKA ADI", "ADI SOYADI", "HESAP NO/ İBAN NO"}.issubset(set(section)):
        return []

    try:
        data_start = section.index("TC NO") + 1
    except ValueError:
        return []

    data = []
    for line in section[data_start:]:
        if line.startswith("5(*)"):
            break
        data.append(line)

    if len(data) < 4:
        return []

    bank_parts = []
    if len(data) >= 3 and data[0] == "YATIRIMCI" and data[1] == "TAZMİN MERKEZİ":
        bank_parts = data[:2]
        data = data[2:]
    else:
        bank_parts = data[:1]
        data = data[1:]

    account_index = next((index for index, line in enumerate(data) if re.search(r"\d", line)), None)
    if account_index is None or account_index < 2:
        return []

    name = normalize_space(" ".join(data[:account_index]))
    account_no = data[account_index]
    amount_hbno = data[account_index + 1] if account_index + 1 < len(data) else ""
    amount_match = re.search(r"\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2}", amount_hbno)
    amount = amount_match.group(0) if amount_match else ""
    hbno = amount_hbno[amount_match.end() :].strip() if amount_match else ""
    identity = data[account_index + 2] if account_index + 2 < len(data) else ""
    identity_digits = normalize_identity(identity)

    return [
        {
            "kisi_ad_soyad_unvan": name,
            "tckn": identity_digits if len(identity_digits) == 11 else defaults.get("keywords_tckn", ""),
            "vkn": identity_digits if len(identity_digits) == 10 else defaults.get("keywords_vkn", ""),
            "haciz_konulan_tutar": amount,
            "hesap_no": account_no,
            "hesap_turu": "TL",
            "hbno": hbno,
            "banka_adi": normalize_space(" ".join(bank_parts)),
            "sube_adi": "",
        }
    ]


def extract_inline_text_records(text: str, defaults: dict[str, str]) -> list[dict[str, str]]:
    flat_text = normalize_space(text)

    identity_match = re.search(
        r"\b(\d{10,11})\s+(?:TC|T\.C\.?|vergi)\s+kimlik\s+numaral[ıi]\s+mükellefi\s+(.+?)(?:'|’|`)?(?:ın|in|un|ün)?\s+vergi\s+borç",
        flat_text,
        flags=re.IGNORECASE,
    )
    if not identity_match:
        return []

    identity = normalize_identity(identity_match.group(1))
    name = normalize_space(identity_match.group(2))

    account_match = re.search(
        r"\bbulunan\s+([A-Z0-9-]{6,})\s+üye\s+iş\s+yeri\s+numaralı",
        flat_text,
        flags=re.IGNORECASE,
    )
    account_no = account_match.group(1) if account_match else ""

    amount = ""
    amount_match = re.search(r"\b\d{1,3}(?:\.\d{3})*,\d{2}\b", flat_text)
    if amount_match:
        amount = amount_match.group(0)

    return [
        {
            "kisi_ad_soyad_unvan": name,
            "tckn": identity if len(identity) == 11 else defaults.get("keywords_tckn", ""),
            "vkn": identity if len(identity) == 10 else defaults.get("keywords_vkn", ""),
            "haciz_konulan_tutar": amount,
            "hesap_no": account_no,
            "hesap_turu": "TL",
            "hbno": "",
            "banka_adi": "",
            "sube_adi": "",
        }
    ]


def find_line_index(lines: list[str], value: str, start: int = 0) -> int | None:
    target = label_key(value)
    for index in range(start, len(lines)):
        if label_key(lines[index]) == target:
            return index
    return None


def value_after_label(lines: list[str], label: str) -> str:
    label_parts = split_label_parts(label)
    normalized_label = normalize_space(" / ".join(label_parts))
    normalized_label_key = label_key(normalized_label)
    for index in range(len(lines)):
        normalized_line = normalize_space(lines[index])
        if label_key(normalized_line).startswith(normalized_label_key):
            tail = normalized_line[len(normalized_label) :].lstrip(":").strip()
            if tail:
                return tail
        consumed = matching_label_length(lines, index, label_parts)
        if consumed:
            return lines[index + consumed] if index + consumed < len(lines) else ""
    return ""


def value_after_label_sequence(lines: list[str], labels: list[str]) -> str:
    keys = [label_key(label) for label in labels]
    for index in range(0, len(lines) - len(labels) + 1):
        if [label_key(line) for line in lines[index : index + len(labels)]] == keys:
            value_index = index + len(labels)
            return lines[value_index] if value_index < len(lines) else ""
    return ""


def split_label_parts(label: str) -> list[str]:
    return [normalize_space(part) for part in label.split(" / ")]


def find_label_line(lines: list[str], label: str) -> str | None:
    normalized_label = normalize_space(label)
    for line in lines:
        if normalize_space(line).startswith(normalized_label):
            return line
    return None


def extract_vkn_tckn_from_text(text: str) -> tuple[str, str]:
    if not text:
        return "", ""
    digits = re.findall(r"\b\d{10,11}\b", text)
    vkn = next((value for value in digits if len(value) == 10), "")
    tckn = next((value for value in digits if len(value) == 11), "")
    return vkn, tckn


def choose_identity_for_entity(name: str, vkn: str, tckn: str) -> str:
    if vkn and tckn:
        return vkn if is_legal_entity(name) else tckn
    return tckn or vkn


def matching_label_length(lines: list[str], index: int, label_parts: list[str]) -> int:
    if index >= len(lines):
        return 0

    if label_key(normalize_space(" / ".join(label_parts))) == label_key(lines[index]):
        return 1

    if len(label_parts) == 2 and index + 2 < len(lines):
        if (
            label_key(lines[index]) == label_key(label_parts[0])
            and lines[index + 1] == "/"
            and label_key(lines[index + 2]) == label_key(label_parts[1])
        ):
            return 3

    if len(label_parts) == 1 and label_key(lines[index]) == label_key(label_parts[0]):
        return 1

    return 0


def clean_amount(value: str) -> str:
    value = normalize_space(value)
    match = re.search(r"\d{1,3}(?:\.\d{3})*,\d{1,2}|\d+,\d{1,2}", value)
    if match:
        amount = match.group(0)
        if re.search(r",\d$", amount):
            amount = f"{amount}0"
        return amount
    digits = re.sub(r"\D", "", value)
    return digits


def extract_table_records(text: str, defaults: dict[str, str]) -> list[dict[str, str]]:
    lines = [normalize_space(line) for line in text.splitlines() if normalize_space(line)]
    records: list[dict[str, str]] = []

    for i, line in enumerate(lines):
        if not re.fullmatch(r"\d{10,11}", line):
            continue

        previous = lines[i - 1] if i > 0 else ""
        next_line = lines[i + 1] if i + 1 < len(lines) else ""
        if len(line) == 11 and re.fullmatch(r"\d{10}", previous):
            continue

        vkn = previous if re.fullmatch(r"\d{10}", previous) else ""
        tckn = line if len(line) == 11 else ""
        if len(line) == 10 and re.fullmatch(r"\d{11}", next_line):
            vkn = line
            tckn = next_line

        if not (vkn or tckn):
            continue

        start = i + 1
        if len(line) == 10 and tckn:
            start = i + 2

        after = lines[start : start + 12]
        name_parts, rest = split_name_and_account(after)
        hbno = find_nearby_hbno(lines, i)
        amount = find_nearby_amount(lines, i) or (extract_amounts(text)[-1] if extract_amounts(text) else "")

        record = {
            "kisi_ad_soyad_unvan": normalize_space(" ".join(name_parts)),
            "tckn": tckn or defaults.get("keywords_tckn", ""),
            "vkn": vkn or defaults.get("keywords_vkn", ""),
            "haciz_konulan_tutar": amount,
            "hesap_no": "",
            "hesap_turu": "TL",
            "hbno": hbno,
            "banka_adi": "",
            "sube_adi": "",
        }
        fill_account_fields(record, rest)

        key = (record["tckn"], record["vkn"], record["hbno"], record["hesap_no"])
        if key not in [
            (r["tckn"], r["vkn"], r["hbno"], r["hesap_no"]) for r in records
        ]:
            records.append(record)

    if not records:
        records.append(
            {
                "kisi_ad_soyad_unvan": "",
                "tckn": defaults.get("keywords_tckn", ""),
                "vkn": defaults.get("keywords_vkn", ""),
                "haciz_konulan_tutar": extract_amounts(text)[-1] if extract_amounts(text) else "",
                "hesap_no": "",
                "hesap_turu": "TL",
                "hbno": "",
                "banka_adi": "",
                "sube_adi": "",
            }
        )

    return records


def extract_positioned_6183_attachment_records(
    pdf_path: Path,
    defaults: dict[str, str],
) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []

    with fitz.open(pdf_path) as doc:
        for page in doc:
            words = [
                word
                for word in page.get_text("words")
                if normalize_space(word[4])
            ]
            if not words:
                continue

            row_markers = sorted(
                (
                    (word[1], word[4])
                    for word in words
                    if 50 <= word[0] <= 85 and re.fullmatch(r"\d+", word[4])
                ),
                key=lambda item: item[0],
            )
            if not row_markers:
                continue

            for index, (row_y, _) in enumerate(row_markers):
                top = row_y - 17.5
                bottom = (
                    row_markers[index + 1][0] - 17.5
                    if index + 1 < len(row_markers)
                    else row_y + 18
                )

                row_words = [
                    word
                    for word in words
                    if top <= word[1] <= bottom and word[1] > 73.8
                ]
                record = positioned_6183_row_record(row_words, defaults)
                if record and any(
                    (
                        record["tckn"],
                        record["vkn"],
                        record["kisi_ad_soyad_unvan"],
                        record["hesap_no"],
                        record["hbno"],
                        record["haciz_konulan_tutar"],
                    )
                ):
                    records.append(record)

    return dedupe_record_dicts(records)


def positioned_6183_row_record(
    words: list[tuple],
    defaults: dict[str, str],
) -> dict[str, str] | None:
    identity = normalize_identity(positioned_text(words, 85, 140))
    if not identity:
        return None

    surname = positioned_text(words, 140, 210)
    name = positioned_text(words, 210, 270)
    hbno = positioned_text(words, 270, 365)
    bank_name = positioned_text(words, 365, 415)
    account_no = positioned_text(words, 415, 465)
    amount = clean_amount(positioned_text(words, 465, 525))

    return {
        "kisi_ad_soyad_unvan": normalize_space(f"{surname} {name}"),
        "tckn": identity if len(identity) == 11 else defaults.get("keywords_tckn", ""),
        "vkn": identity if len(identity) == 10 else defaults.get("keywords_vkn", ""),
        "haciz_konulan_tutar": amount,
        "hesap_no": account_no,
        "hesap_turu": "TL",
        "hbno": hbno,
        "banka_adi": bank_name,
        "sube_adi": "",
    }


def positioned_text(words: list[tuple], left: float, right: float) -> str:
    column_words = [
        word
        for word in words
        if left <= word[0] < right and not is_positioned_header_word(word)
    ]
    column_words.sort(key=lambda word: (round(word[1], 1), word[0]))
    return normalize_space(" ".join(word[4] for word in column_words))


def is_positioned_header_word(word: tuple) -> bool:
    text = label_key(word[4])
    if word[1] > 74:
        return False
    return text in {
        "sira",
        "no",
        "tc",
        "kimlik",
        "mukellef",
        "soyadi",
        "adi",
        "hb",
        "numarasi",
        "banka",
        "hesap",
        "tatbik",
        "tutar",
        "haciz",
    }


def dedupe_record_dicts(records: list[dict[str, str]]) -> list[dict[str, str]]:
    deduped: list[dict[str, str]] = []
    seen: set[tuple[str, str, str, str, str, str]] = set()
    for record in records:
        key = (
            record.get("tckn", ""),
            record.get("vkn", ""),
            record.get("kisi_ad_soyad_unvan", ""),
            record.get("hbno", ""),
            record.get("hesap_no", ""),
            record.get("haciz_konulan_tutar", ""),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(record)
    return deduped


def split_name_and_account(lines: list[str]) -> tuple[list[str], list[str]]:
    name_parts: list[str] = []
    account_start = len(lines)
    stop_words = {
        "Merkezi Kayıt",
        "Kuruluşu A.Ş",
        "YATIRIMCI",
        "TAZMİN",
        "MERKEZİ",
    }

    for index, line in enumerate(lines):
        if (
            line in stop_words
            or re.search(r"\d", line)
            or "BANK" in line.upper()
            or "ŞUBE" in line.upper()
        ):
            account_start = index
            break
        if re.match(r"^[A-ZÇĞİÖŞÜ0-9 .,&()/-]+$", line):
            name_parts.append(line)
        else:
            account_start = index
            break

    return name_parts, lines[account_start:]


def find_nearby_hbno(lines: list[str], index: int) -> str:
    for line in reversed(lines[max(0, index - 5) : index]):
        if re.search(r"[A-Za-z]", line) and re.search(r"\d", line) and len(line) >= 12:
            return line
    return ""


def find_nearby_amount(lines: list[str], index: int) -> str:
    for line in lines[index : index + 18]:
        match = re.search(r"\b\d{1,3}(?:\.\d{3})*,\d{2}\b", line)
        if match:
            return match.group(0)
    return ""


def fill_account_fields(record: dict[str, str], lines: list[str]) -> None:
    cleaned = []
    for line in lines:
        if re.fullmatch(r"\d{1,3}(?:\.\d{3})*,\d{2}", line):
            continue
        if line.startswith("Nurgül ") or line.startswith("Bu belge"):
            break
        cleaned.append(line)

    if cleaned:
        record["hesap_no"] = normalize_space(" ".join(cleaned))

    hesap_no = ""
    for line in cleaned:
        if re.search(r"\d", line) and not re.fullmatch(r"\d{10,11}", line):
            hesap_no = line
            break

    if hesap_no:
        before_hesap = cleaned[: cleaned.index(hesap_no)]
    else:
        before_hesap = cleaned

    if before_hesap:
        midpoint = max(1, len(before_hesap) // 2)
        record["banka_adi"] = normalize_space(" ".join(before_hesap[:midpoint]))
        record["sube_adi"] = normalize_space(" ".join(before_hesap[midpoint:]))
    if hesap_no:
        record["hesap_no"] = hesap_no


def is_attachment_pdf(pdf_path: Path) -> bool:
    return pdf_path.name.casefold().endswith(".ek.pdf")


def attachment_pdf_for(pdf_path: Path) -> Path | None:
    if is_attachment_pdf(pdf_path):
        return None
    attachment_path = pdf_path.with_name(f"{pdf_path.stem}.ek{pdf_path.suffix}")
    return attachment_path if attachment_path.exists() else None


def parse_pdf(pdf_path: Path, taxpayer_pdf_path: Path | None = None) -> list[EypRecord]:
    defaults = {
        "gonderen_kurum": "",
        "sayi": "",
        "tarih": "",
        "konu": "",
        "keywords_tckn": "",
        "keywords_vkn": "",
    }
    text = pdf_text(pdf_path)
    header = get_pdf_header_fields(text)
    taxpayer_path = taxpayer_pdf_path or pdf_path
    taxpayer_text = text if taxpayer_path == pdf_path else pdf_text(taxpayer_path)

    for key in ("gonderen_kurum", "sayi", "tarih"):
        if header.get(key):
            defaults[key] = header[key]

    table_records = extract_taxpayer_text_grid_records(taxpayer_text, defaults)
    if not table_records:
        table_records = pdf_table_records(taxpayer_path, defaults)
    if not table_records:
        table_records = extract_vertical_taxpayer_records(taxpayer_text, defaults)
    if not table_records:
        table_records = extract_horizontal_records(taxpayer_text, defaults)
    if not table_records:
        table_records = extract_inline_text_records(taxpayer_text, defaults)
    if not table_records:
        table_records = extract_court_records(taxpayer_text, defaults)
    if not table_records:
        table_records = extract_bank_transfer_grid_records(taxpayer_path, defaults)
    if not table_records:
        table_records = extract_positioned_6183_attachment_records(taxpayer_path, defaults)
    if not table_records:
        table_records = extract_table_records(taxpayer_text, defaults)
    records = []
    for record in table_records:
        output_record = record.copy()
        output_record["kisi_ad_soyad_unvan"] = format_name_or_title(
            output_record.get("kisi_ad_soyad_unvan", "")
        )
        output_record["hesap_turu"] = output_record.get("hesap_turu") or "TL"
        output_record["haciz_konulan_tutar"] = clean_amount(
            output_record.get("haciz_konulan_tutar", "")
        )
        output_record["kisi_turu"] = person_type_for_record(output_record)
        output_record["tckn_vkn"] = identity_number_for_record(output_record)
        output_record.pop("tckn", None)
        output_record.pop("vkn", None)
        output_record.pop("banka_adi", None)
        output_record.pop("sube_adi", None)
        records.append(
            EypRecord(
                gonderen_kurum=defaults.get("gonderen_kurum", ""),
                sayi=defaults.get("sayi", ""),
                tarih=defaults.get("tarih", ""),
                pdf_dosyasi=pdf_path.name,
                **output_record,
            )
        )
    return records


def extract_court_records(text: str, defaults: dict[str, str]) -> list[dict[str, str]]:
    lines = [normalize_space(line) for line in text.splitlines() if normalize_space(line)]
    records = []
    seen_identities: set[tuple[str, str, str]] = set()

    for i, line in enumerate(lines):
        if label_key(line).replace(" ", "") != "ilgilikisi":
            continue
        context = court_related_person_context(lines, i)
        name, tckn, vkn = parse_court_person_context(context)
        if name or tckn or vkn:
            key = (name, tckn, vkn)
            if key in seen_identities:
                continue
            seen_identities.add(key)
            records.append({
                "kisi_ad_soyad_unvan": name,
                "tckn": tckn or defaults.get("keywords_tckn", ""),
                "vkn": vkn,
                "haciz_konulan_tutar": "",
                "hesap_no": "",
                "hesap_turu": "TL",
                "hbno": "",
                "banka_adi": "",
                "sube_adi": "",
            })
    if records:
        return records

    for i, line in enumerate(lines):
        normalized_line = normalize_text(line)
        if "muris" in normalized_line or "muteveffa" in normalized_line or "muflis" in normalized_line:
            context = court_person_context(lines, i)
            name, tckn, vkn = parse_court_person_context(context)
            if name or tckn or vkn:
                key = (name, tckn, vkn)
                if key in seen_identities:
                    continue
                seen_identities.add(key)
                records.append({
                    "kisi_ad_soyad_unvan": name,
                    "tckn": tckn or defaults.get("keywords_tckn", ""),
                    "vkn": vkn or defaults.get("keywords_vkn", ""),
                    "haciz_konulan_tutar": "",
                    "hesap_no": "",
                    "hesap_turu": "TL",
                    "hbno": "",
                    "banka_adi": "",
                    "sube_adi": "",
                })

    return records


def court_related_person_context(lines: list[str], index: int) -> str:
    context_lines = [lines[index]]
    for line in lines[index + 1 : min(index + 4, len(lines))]:
        context_lines.append(line)
        if re.search(r"\b\d{10,11}\b", line):
            break
    return normalize_space(" ".join(context_lines))


def court_person_context(lines: list[str], index: int) -> str:
    context_lines = [lines[index]]
    stop_markers = {
        "dagitim listesi",
        "ek :",
        "ek:",
        "katip",
        "hakim",
        "e-imzalidir",
        "bilgilerinize",
    }
    for line in lines[index + 1 : min(index + 6, len(lines))]:
        normalized = normalize_text(line)
        if any(normalized.startswith(marker) for marker in stop_markers):
            break
        context_lines.append(line)
        if re.search(r"\b\d{10,11}\b", line):
            break
    return normalize_space(" ".join(context_lines))


def parse_court_person_context(context: str) -> tuple[str, str, str]:
    digits = re.findall(r"\b\d{10,11}\b", context)
    tckn = next((value for value in digits if len(value) == 11), "")
    vkn = next((value for value in digits if len(value) == 10), "")
    identity = tckn or vkn
    if not identity:
        return clean_court_person_name(context), "", ""

    before_identity = context.split(identity, 1)[0]
    name = clean_court_person_name(before_identity)
    if not name:
        after_identity = context.split(identity, 1)[1]
        name = clean_court_person_name(after_identity)
    return name, tckn, vkn


def clean_court_person_name(value: str) -> str:
    value = normalize_space(value)
    value = re.sub(
        r"(?i)\b(?:ilgili\s+kişi|ilgili\s+kisi|muris|m[üu]teveffa|m[üu]flis)\b\s*:?",
        " ",
        value,
    )
    value = re.sub(r"\([^\)]*$", " ", value)
    value = re.sub(r"[-:]+$", " ", value)
    value = normalize_space(value)
    value = re.sub(r"(?:'|`|’)[^\s]+$", " ", value)
    value = re.sub(r"(?:'|`|’)?(?:n[ıiuü]n|in|[ıiuü]n|un|ün)$", " ", value, flags=re.IGNORECASE)
    value = re.split(
        r"\b(?:ili|il[çc]esi|mah/k[öo]y|cilt|aile|s[ıi]ra|n[üu]fusa|terekesinin)\b",
        value,
        maxsplit=1,
        flags=re.IGNORECASE,
    )[0]
    value = re.sub(r"\b\d{10,11}\b", " ", value)
    value = re.sub(r"[^\w\s.'-]", " ", value, flags=re.UNICODE)
    value = re.sub(r"[\d_]", " ", value)
    value = normalize_space(value).strip(" -:;,.()")
    return value


def amount_to_float(value) -> float:
    parsed = parse_numeric_amount(value)
    if isinstance(parsed, (int, float)):
        return float(parsed)
    try:
        normalized = normalize_space(str(parsed)).replace(".", "").replace(",", ".")
        return float(normalized)
    except ValueError:
        return 0.0


def merge_pdf_records(records: list[EypRecord]) -> list[EypRecord]:
    merged: dict[tuple[str, str, str, str], EypRecord] = {}
    for record in records:
        key = (
            normalize_space(record.tckn_vkn),
            normalize_space(record.kisi_ad_soyad_unvan),
            normalize_space(record.hesap_no),
            normalize_space(record.hbno),
        )
        amount = amount_to_float(record.haciz_konulan_tutar)
        record.haciz_konulan_tutar = amount
        if key in merged:
            merged_record = merged[key]
            merged_record.haciz_konulan_tutar = amount_to_float(
                merged_record.haciz_konulan_tutar
            ) + amount
        else:
            merged[key] = record
    return list(merged.values())


def load_balance_data(csv_path: Path) -> tuple[dict[str, str], dict[str, set[str]], set[str]]:
    if not csv_path.exists():
        return {}, {}, set()

    balance_lookup: dict[str, str] = {}
    account_lookup: dict[str, set[str]] = {}
    all_accounts: set[str] = set()
    for encoding in ("utf-8-sig", "cp1254"):
        try:
            with csv_path.open("r", encoding=encoding, newline="") as csv_file:
                reader = csv.reader(csv_file)
                next(reader, None)
                for row in reader:
                    if len(row) < 6:
                        continue
                    identity = normalize_identity(row[3])
                    if not identity:
                        continue
                    if identity not in balance_lookup:
                        balance_lookup[identity] = f"{normalize_space(row[4])}-%{normalize_space(row[5])}"
                    if len(row) >= 12:
                        account_no = normalize_account_no(row[11])
                        if account_no:
                            account_lookup.setdefault(identity, set()).add(account_no)
                            all_accounts.add(account_no)
            return balance_lookup, account_lookup, all_accounts
        except UnicodeDecodeError:
            balance_lookup.clear()
            account_lookup.clear()
            all_accounts.clear()
            continue
    return balance_lookup, account_lookup, all_accounts


def normalize_identity(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def normalize_account_no(value: str) -> str:
    return re.sub(r"\s+", "", value or "").upper()


def account_no_candidates(value: str) -> set[str]:
    normalized = normalize_account_no(value)
    candidates = {normalized} if normalized else set()
    searchable = (value or "").upper()
    candidates.update(
        token
        for token in re.findall(r"[A-Z0-9]+", searchable)
        if len(token) >= 4 and re.search(r"\d", token)
    )
    return candidates


def fill_balance_info(
    records: list[EypRecord],
    balance_lookup: dict[str, str],
    account_lookup: dict[str, set[str]],
    all_accounts: set[str],
) -> None:
    for record in records:
        identity = normalize_identity(record.tckn_vkn)
        if not normalize_account_no(record.hesap_no):
            record.hesap_no = account_no_for_identity(identity, account_lookup)
        record.HESAP_TURU = balance_lookup.get(identity, "")
        record.hesap_kontrol = account_check_result(
            identity, record.hesap_no, account_lookup, all_accounts
        )


def account_no_for_identity(identity: str, account_lookup: dict[str, set[str]]) -> str:
    accounts = account_lookup.get(identity, set())
    return " / ".join(sorted(accounts)) if accounts else ""


def account_check_result(
    identity: str,
    account_no: str,
    account_lookup: dict[str, set[str]],
    all_accounts: set[str],
) -> str:
    candidates = account_no_candidates(account_no)
    if identity not in account_lookup:
        return matching_account_no(candidates, all_accounts) or "Yok"
    if not normalize_account_no(account_no):
        return first_account_no(account_lookup[identity])
    if candidates & account_lookup[identity]:
        return "Doğru"
    return "Hatalı"


def matching_account_no(candidates: set[str], accounts: set[str]) -> str:
    matches = candidates & accounts
    return sorted(matches)[0] if matches else ""


def first_account_no(accounts: set[str]) -> str:
    return sorted(accounts)[0] if accounts else "Yok"


def process_folder(folder: Path, output: Path) -> list[EypRecord]:
    pdf_folder = folder / "yazilar"
    if not pdf_folder.exists():
        raise FileNotFoundError(f"{folder} içinde yazilar alt klasörü bulunamadı.")
    balance_lookup, account_lookup, all_accounts = load_balance_data(
        folder / "bakiye_data" / "rpYtmBakiyeRapor.csv"
    )

    workbook = open_or_create_workbook(output)
    ensure_sheet(workbook, SHEET_6183)
    ensure_sheet(workbook, SHEET_TEREKE)
    processed_pdfs = existing_pdf_names(workbook)
    new_pdf_names: set[str] = set()
    all_records: list[EypRecord] = []
    for pdf_path in sorted(pdf_folder.glob("*.pdf")):
        if is_attachment_pdf(pdf_path):
            continue
        if pdf_path.name in processed_pdfs:
            continue
        records = merge_pdf_records(parse_pdf(pdf_path, attachment_pdf_for(pdf_path)))
        fill_balance_info(records, balance_lookup, account_lookup, all_accounts)
        appended_records = append_records_to_workbook(
            workbook,
            records,
            pdf_folder,
        )
        if appended_records:
            new_pdf_names.add(pdf_path.name)
        all_records.extend(appended_records)

    if not any(pdf_folder.glob("*.pdf")):
        raise FileNotFoundError(f"{pdf_folder} içinde .pdf dosyası bulunamadı.")

    format_workbook(workbook, pdf_folder, balance_lookup, account_lookup, all_accounts)
    highlight_rows_by_pdf_names(workbook, new_pdf_names)
    workbook.save(output)
    return all_records


def open_or_create_workbook(output: Path):
    if output.exists():
        workbook = load_workbook(output)
        migrate_existing_result_sheet(workbook)
        return workbook

    workbook = Workbook()
    workbook.active.title = SHEET_6183
    return workbook


def migrate_existing_result_sheet(workbook) -> None:
    if SHEET_6183 in workbook.sheetnames:
        return
    for sheet in workbook.worksheets:
        if header_row_index(sheet):
            sheet.title = SHEET_6183
            return


def ensure_sheet(workbook, title: str):
    if title in workbook.sheetnames:
        sheet = workbook[title]
    else:
        sheet = workbook.create_sheet(title)

    normalize_sheet_headers(sheet)
    return sheet


def normalize_sheet_headers(sheet) -> None:
    if all(value is None for value in sheet_row_values(sheet, 1)) and header_map(sheet, 2):
        sheet.delete_rows(1)

    desired_columns = columns_for_sheet(sheet)
    if sheet_row_values(sheet, 1) == desired_columns:
        return

    current_header_map = header_map(sheet, 1)
    if current_header_map:
        migrate_sheet_layout(sheet, current_header_map, desired_columns)
        return

    if all(value is None for value in sheet_row_values(sheet, 1)):
        write_headers(sheet, desired_columns)


def columns_for_sheet(sheet) -> list[str | None]:
    if sheet.title == SHEET_6183:
        return SHEET_6183_COLUMNS
    return OUTPUT_COLUMNS


def sheet_row_values(sheet, row: int) -> list[str | None]:
    width = max(len(columns_for_sheet(sheet)), sheet.max_column)
    return [sheet.cell(row=row, column=column).value for column in range(1, width + 1)]


def write_headers(sheet, columns: list[str | None]) -> None:
    for column_index, header in enumerate(columns, start=1):
        sheet.cell(row=1, column=column_index).value = header


def header_map(sheet, row: int) -> dict[str, int]:
    values = sheet_row_values(sheet, row)
    mapping = {
        str(value): column_index
        for column_index, value in enumerate(values, start=1)
        if value in INTERNAL_COLUMNS
    }
    if all(column in mapping for column in OUTPUT_COLUMNS):
        return mapping
    return {}


def migrate_sheet_layout(sheet, current_header_map: dict[str, int], desired_columns: list[str | None]) -> None:
    rows = []
    for row in range(2, sheet.max_row + 1):
        row_data = {}
        for column, column_index in current_header_map.items():
            row_data[column] = sheet.cell(row=row, column=column_index).value
        if any(value not in (None, "") for value in row_data.values()):
            rows.append(row_data)

    sheet.delete_rows(1, sheet.max_row)
    write_headers(sheet, desired_columns)
    for row_data in rows:
        append_record_dict_to_sheet(sheet, row_data)


def header_row_index(sheet) -> int | None:
    for row in range(1, min(sheet.max_row, 5) + 1):
        if header_map(sheet, row):
            return row
    return None


def existing_pdf_names(workbook) -> set[str]:
    names: set[str] = set()
    for sheet in workbook.worksheets:
        header_row = header_row_index(sheet)
        if not header_row:
            continue
        headers = header_map(sheet, header_row)
        pdf_column = headers["pdf_dosyasi"]
        for row in range(header_row + 1, sheet.max_row + 1):
            value = sheet.cell(row=row, column=pdf_column).value
            if value:
                names.add(str(value))
    return names


def remove_pdf_rows(workbook, pdf_name: str) -> dict[tuple[str, ...], str]:
    preserved_a_values: dict[tuple[str, ...], str] = {}
    for sheet in workbook.worksheets:
        header_row = header_row_index(sheet)
        if not header_row:
            continue
        headers = header_map(sheet, header_row)
        if "pdf_dosyasi" not in headers:
            continue
        pdf_column = headers["pdf_dosyasi"]
        for row in range(sheet.max_row, header_row, -1):
            if sheet.cell(row=row, column=pdf_column).value == pdf_name:
                a_value = sheet.cell(row=row, column=1).value
                if normalize_space(str(a_value or "")).lower() == "ok":
                    continue
                if a_value not in (None, ""):
                    row_key = row_key_from_sheet_row(sheet, row, headers)
                    preserved_a_values[row_key] = str(a_value)
                sheet.delete_rows(row)
    return preserved_a_values


NEW_RECORD_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")


def append_records_to_workbook(
    workbook,
    records: list[EypRecord],
    pdf_folder: Path,
    preserved_a_values: dict[tuple[str, ...], str] | None = None,
) -> list[EypRecord]:
    preserved_a_values = preserved_a_values or {}
    existing_rows = existing_record_rows(workbook)
    appended_records: list[EypRecord] = []
    for record in records:
        row_key = record_row_key(record)
        if row_key in existing_rows:
            continue
        sheet = workbook[sheet_name_for_record(record)]
        append_record_dict_to_sheet(
            sheet,
            record.as_dict(),
            preserved_a_values.get(row_key, ""),
        )
        format_row(sheet, sheet.max_row, pdf_folder)
        existing_rows.add(row_key)
        appended_records.append(record)
    return appended_records


def parse_numeric_amount(value) -> float | str:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        return value

    value = normalize_space(str(value))
    if not value:
        return ""

    match = re.search(r"\d{1,3}(?:[\.,]\d{3})*(?:[\.,]\d+)?", value)
    if not match:
        return value

    numeric = match.group(0)
    if "," in numeric and "." in numeric:
        numeric = numeric.replace(".", "").replace(",", ".")
    elif "," in numeric:
        numeric = numeric.replace(",", ".")
    elif "." in numeric and numeric.count(".") > 1:
        numeric = numeric.replace(".", "")

    try:
        return float(numeric)
    except ValueError:
        return value


def append_record_dict_to_sheet(
    sheet,
    values: dict[str, str],
    preserved_a_value: str = "",
) -> None:
    row = ["" for _ in columns_for_sheet(sheet)]
    legal_entity = is_legal_entity(values.get("kisi_ad_soyad_unvan", ""))
    for index, column in enumerate(columns_for_sheet(sheet)):
        if not column:
            continue
        cell_value = values.get(column, values.get("bakiye_bilgi", ""))
        if column == "haciz_konulan_tutar":
            cell_value = parse_numeric_amount(cell_value)
        if column == "gonderen_kurum":
            cell_value = format_sender_for_sheet(sheet.title, cell_value)
        if legal_entity and column == "sayi":
            cell_value = format_legal_entity_name(str(cell_value))
        row[index] = cell_value

    if preserved_a_value and columns_for_sheet(sheet) and columns_for_sheet(sheet)[0] is None:
        row[0] = preserved_a_value

    sheet.append(row)


def existing_record_rows(workbook) -> set[tuple[str, ...]]:
    rows: set[tuple[str, ...]] = set()
    for sheet in workbook.worksheets:
        header_row = header_row_index(sheet)
        if not header_row:
            continue
        headers = header_map(sheet, header_row)
        column_indexes = [headers[column] for column in OUTPUT_COLUMNS]
        for row in range(header_row + 1, sheet.max_row + 1):
            values = [sheet.cell(row=row, column=column_index).value for column_index in column_indexes]
            if any(value not in (None, "") for value in values):
                rows.add(tuple(normalize_cell_value(value) for value in values))
    return rows


def record_row_key(record: EypRecord) -> tuple[str, ...]:
    values = record.as_dict()
    return tuple(normalize_cell_value(values.get(column, "")) for column in OUTPUT_COLUMNS)


def row_key_from_sheet_row(sheet, row: int, headers: dict[str, int]) -> tuple[str, ...]:
    values = [sheet.cell(row=row, column=headers[column]).value for column in OUTPUT_COLUMNS]
    return tuple(normalize_cell_value(value) for value in values)


def normalize_cell_value(value) -> str:
    return normalize_space(str(value)) if value is not None else ""


def highlight_rows_by_pdf_names(workbook, pdf_names: set[str]) -> None:
    if not pdf_names:
        return
    for sheet in workbook.worksheets:
        header_row = header_row_index(sheet)
        if not header_row:
            continue
        headers = header_map(sheet, header_row)
        if "pdf_dosyasi" not in headers:
            continue
        pdf_column = headers["pdf_dosyasi"]
        for row in range(header_row + 1, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=pdf_column)
            if str(cell.value or "") in pdf_names:
                for column_index in range(1, sheet.max_column + 1):
                    target = sheet.cell(row=row, column=column_index)
                    if target.value not in (None, ""):
                        target.fill = NEW_RECORD_FILL


def sheet_name_for_record(record: EypRecord) -> str:
    if is_court_sender(record.gonderen_kurum):
        return SHEET_TEREKE
    return SHEET_6183


def normalize_text(text: str) -> str:
    text = text.replace('I', 'i')
    text = text.replace('İ', 'i')
    text = text.replace('ı', 'i')
    text = text.replace('Ğ', 'g')
    text = text.replace('ğ', 'g')
    text = text.replace('Ü', 'u')
    text = text.replace('ü', 'u')
    text = text.replace('Ş', 's')
    text = text.replace('ş', 's')
    text = text.replace('Ö', 'o')
    text = text.replace('ö', 'o')
    text = text.replace('Ç', 'c')
    text = text.replace('ç', 'c')
    text = text.lower()
    return text


def is_court_sender(sender: str) -> bool:
    normalized = normalize_text(sender)
    court_indicators = [
        "mahkeme",
        "adliye",
        "satis mudurlugu",
        "icra dairesi",
        "icra mudurlugu"
    ]
    return any(indicator in normalized for indicator in court_indicators)


def format_workbook(
    workbook,
    pdf_folder: Path,
    balance_lookup: dict[str, str],
    account_lookup: dict[str, set[str]],
    all_accounts: set[str],
) -> None:
    for sheet_name in (SHEET_6183, SHEET_TEREKE):
        sheet = ensure_sheet(workbook, sheet_name)
        sort_sheet_by_date(sheet)
        header_row = header_row_index(sheet)
        if not header_row:
            continue
        fill_account_no_on_sheet(sheet, header_row, account_lookup)
        if sheet.title == SHEET_6183:
            fill_balance_info_on_sheet(sheet, header_row, balance_lookup)
            fill_account_check_on_sheet(sheet, header_row, account_lookup, all_accounts)
        for row in range(header_row + 1, sheet.max_row + 1):
            format_row(sheet, row, pdf_folder)


def fill_balance_info_on_sheet(sheet, header_row: int, balance_lookup: dict[str, str]) -> None:
    headers = header_map(sheet, header_row)
    balance_header = BALANCE_INFO_COLUMN if BALANCE_INFO_COLUMN in headers else "bakiye_bilgi"
    if "tckn_vkn" not in headers or balance_header not in headers:
        return
    identity_column = headers["tckn_vkn"]
    balance_column = headers[balance_header]
    for row in range(header_row + 1, sheet.max_row + 1):
        identity = normalize_identity(str(sheet.cell(row=row, column=identity_column).value or ""))
        sheet.cell(row=row, column=balance_column).value = balance_lookup.get(identity, "")


def fill_account_no_on_sheet(sheet, header_row: int, account_lookup: dict[str, set[str]]) -> None:
    headers = header_map(sheet, header_row)
    if "tckn_vkn" not in headers or "hesap_no" not in headers:
        return
    identity_column = headers["tckn_vkn"]
    account_column = headers["hesap_no"]
    for row in range(header_row + 1, sheet.max_row + 1):
        current_account = str(sheet.cell(row=row, column=account_column).value or "")
        if normalize_account_no(current_account):
            continue
        identity = normalize_identity(str(sheet.cell(row=row, column=identity_column).value or ""))
        account_no = account_no_for_identity(identity, account_lookup)
        if account_no:
            sheet.cell(row=row, column=account_column).value = account_no


def fill_account_check_on_sheet(
    sheet, header_row: int, account_lookup: dict[str, set[str]], all_accounts: set[str]
) -> None:
    headers = header_map(sheet, header_row)
    if "tckn_vkn" not in headers or "hesap_no" not in headers or ACCOUNT_CHECK_COLUMN not in headers:
        return
    identity_column = headers["tckn_vkn"]
    account_column = headers["hesap_no"]
    result_column = headers[ACCOUNT_CHECK_COLUMN]
    for row in range(header_row + 1, sheet.max_row + 1):
        identity = normalize_identity(str(sheet.cell(row=row, column=identity_column).value or ""))
        account_no = str(sheet.cell(row=row, column=account_column).value or "")
        sheet.cell(row=row, column=result_column).value = account_check_result(
            identity, account_no, account_lookup, all_accounts
        )


def sort_sheet_by_date(sheet) -> None:
    header_row = header_row_index(sheet)
    if not header_row or sheet.max_row <= header_row + 1:
        return

    headers = header_map(sheet, header_row)

    data_rows = [
        row_dict_from_sheet(sheet, row, headers)
        for row in range(header_row + 1, sheet.max_row + 1)
        if any(sheet.cell(row=row, column=column_index).value not in (None, "") for column_index in headers.values())
    ]
    data_rows.sort(key=lambda values: parse_excel_date(values.get("tarih")))

    if sheet.max_row > header_row:
        sheet.delete_rows(header_row + 1, sheet.max_row - header_row)
    for values in data_rows:
        append_record_dict_to_sheet(sheet, values)


def row_dict_from_sheet(sheet, row: int, headers: dict[str, int]) -> dict[str, str]:
    result = {}
    for column, column_index in headers.items():
        value = sheet.cell(row=row, column=column_index).value
        if column == "tarih" and isinstance(value, datetime):
            value = value.strftime("%d.%m.%Y")
        result[column] = value
    return result


def parse_excel_date(value) -> datetime:
    if isinstance(value, datetime):
        return value
    value = normalize_space(str(value)) if value is not None else ""
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return datetime.max


def format_row(sheet, row: int, pdf_folder: Path) -> None:
    text_columns = {
        "sayi",
        "tarih",
        "tckn_vkn",
        "hbno",
        "hesap_no",
        BALANCE_INFO_COLUMN,
        ACCOUNT_CHECK_COLUMN,
    }
    header_row = header_row_index(sheet) or 1
    headers = header_map(sheet, header_row)
    for header, column_index in headers.items():
        if header in text_columns:
            sheet.cell(row=row, column=column_index).number_format = "@"
        if header == "haciz_konulan_tutar":
            sheet.cell(row=row, column=column_index).number_format = "#,##0.00"
        if header == "gonderen_kurum":
            cell = sheet.cell(row=row, column=column_index)
            cell.value = format_sender_for_sheet(sheet.title, cell.value)
            cell.alignment = Alignment(wrap_text=True)
        if header == "pdf_dosyasi":
            cell = sheet.cell(row=row, column=column_index)
            if not cell.value:
                continue
            pdf_path = (pdf_folder / str(cell.value)).resolve()
            cell.hyperlink = str(pdf_path)
            cell.font = Font(color="0563C1", underline="single")


def main() -> None:
    parser = argparse.ArgumentParser(
        description=f"{READING_6183} ve {READING_TEREKE} sonuçlarını Excel'e aktarır."
    )
    parser.add_argument(
        "folder",
        nargs="?",
        default=None,
        help="PDF dosyalarının bulunduğu klasör. Varsayılan: script klasörü.",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="pdf_sonuclar.xlsx",
        help="Oluşturulacak Excel dosyası. Varsayılan: pdf_sonuclar.xlsx",
    )
    args = parser.parse_args()

    folder = Path(args.folder).resolve() if args.folder else Path(__file__).resolve().parent
    output_arg = Path(args.output)
    output = output_arg.resolve() if output_arg.is_absolute() else (folder / output_arg).resolve()
    records = process_folder(folder, output)
    count_6183 = sum(1 for record in records if sheet_name_for_record(record) == SHEET_6183)
    count_tereke = sum(1 for record in records if sheet_name_for_record(record) == SHEET_TEREKE)
    print(f"{READING_6183}: {count_6183} yeni kayıt eklendi.")
    print(f"{READING_TEREKE}: {count_tereke} yeni kayıt eklendi.")
    print(f"Toplam: {len(records)} yeni kayıt eklendi.")
    print(f"Excel: {output}")


if __name__ == "__main__":
    main()
