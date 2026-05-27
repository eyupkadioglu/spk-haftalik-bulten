import io
import os
import uuid
from datetime import datetime, date
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file, current_app
from flask_login import login_required, current_user
from app.extensions import db
from app.models.cari import Cari
from app.models.stok import Stok, StokKategori, StokHareket
from app.models.kasa import KasaHesap, KasaHareket
from app.utils import next_sequence

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

bp = Blueprint('import_export', __name__)

ALLOWED_EXT = {'xlsx', 'xls', 'csv'}


def _ext_ok(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT


def _read_workbook(file_storage):
    """Return list of dicts from first sheet. Handles xlsx/csv."""
    name = file_storage.filename.lower()
    data = file_storage.read()
    if name.endswith('.csv'):
        import csv
        rows = list(csv.DictReader(io.StringIO(data.decode('utf-8-sig'))))
        return rows
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
    return rows


def _find(row, *keys):
    """Try multiple column name variants (exact, then startswith, then contains)."""
    normalized = {col.lower().strip(): val for col, val in row.items()}
    # 1. Exact match
    for k in keys:
        kl = k.lower()
        if kl in normalized:
            return normalized[kl]
    # 2. Starts-with match (handles "Alış Para Birimi (TRY/USD/EUR)" when key is "alış para birimi")
    for k in keys:
        kl = k.lower()
        for col, val in normalized.items():
            if col.startswith(kl):
                return val
    # 3. Contains match (last resort)
    for k in keys:
        kl = k.lower()
        for col, val in normalized.items():
            if kl in col:
                return val
    return None


def _to_float(v, default=0.0):
    if v is None:
        return default
    # Numeric value from openpyxl — use directly
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(' ', '').replace('₺', '').replace('TL', '').replace('$', '').replace('€', '')
    if not s:
        return default
    # Detect Turkish locale format:
    # "1.234,56"  → 1234.56  (period=thousands, comma=decimal)
    # "2.000"     → 2000     (period=thousands, no decimal part)
    # "2,5"       → 2.5      (comma=decimal, no thousands)
    # "1234.56"   → 1234.56  (standard float, period=decimal)
    has_dot = '.' in s
    has_comma = ',' in s
    try:
        if has_comma and has_dot:
            # e.g. "1.234,56" — comma is last, so it's decimal
            last_dot = s.rfind('.')
            last_comma = s.rfind(',')
            if last_comma > last_dot:
                # Turkish: periods=thousands, comma=decimal
                s = s.replace('.', '').replace(',', '.')
            else:
                # English: commas=thousands, period=decimal
                s = s.replace(',', '')
        elif has_comma and not has_dot:
            # e.g. "2,5" or "1.234" — comma likely decimal
            # If digits after comma == 3, could be thousands but assume decimal
            after_comma = s.rsplit(',', 1)[-1]
            s = s.replace(',', '.')
        elif has_dot and not has_comma:
            dot_count = s.count('.')
            if dot_count > 1:
                # Multiple dots → Turkish thousands: "1.234.567" → 1234567
                s = s.replace('.', '')
            # single dot → treat as decimal: "2.5" → 2.5, "2.000" → 2.0
        return float(s)
    except (ValueError, TypeError):
        return default


def _to_date(v):
    if v is None:
        return date.today()
    if isinstance(v, (date, datetime)):
        return v.date() if isinstance(v, datetime) else v
    s = str(v).strip()
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return date.today()


# ─── Ana sayfa ───────────────────────────────────────────────────────────────

@bp.route('/')
@login_required
def index():
    kasa_hesaplari = KasaHesap.query.filter_by(is_active=True).all()
    return render_template('import_export/index.html', kasa_hesaplari=kasa_hesaplari)


# ─── Şablon indir ─────────────────────────────────────────────────────────────

@bp.route('/sablon/<tip>')
@login_required
def sablon_indir(tip):
    if not HAS_OPENPYXL:
        flash('openpyxl kütüphanesi yüklü değil.', 'danger')
        return redirect(url_for('import_export.index'))

    sablonlar = {
        'cari': {
            'filename': 'sablon_cari.xlsx',
            'headers': ['Cari Kodu', 'Ünvan', 'Cari Tipi (musteri/tedarikci/her_ikisi)',
                        'Vergi No', 'TCKN', 'Vergi Dairesi', 'Telefon', 'E-posta',
                        'Adres', 'Şehir', 'Ülke', 'Açık Hesap Limiti', 'Ödeme Vadesi (gün)', 'Notlar'],
            'sample': ['C-001', 'Örnek Müşteri A.Ş.', 'musteri', '1234567890', '',
                       'Büyük Mükellefler', '0212 000 0000', 'info@ornek.com',
                       'İstanbul Cad. No:1', 'İstanbul', 'Türkiye', '10000', '30', ''],
        },
        'stok': {
            'filename': 'sablon_stok.xlsx',
            'headers': [
                'Stok Kodu', 'Barkod', 'Ürün Adı', 'Ürün Tipi',
                'Marka', 'Kategori', 'Birim',
                'Alış Para Birimi', 'Alış Fiyatı', 'Alış KDV Oranı',
                'Satış Para Birimi', 'Satış Fiyatı', 'Satış KDV Oranı',
                'Tedarikçi Kodu', 'Mevcut Stok', 'Min Stok', 'Açıklama',
            ],
            'sample': [
                'STK-001', '8690000000001', 'Örnek Ürün', 'stok',
                'MarkA', 'Genel', 'Adet',
                'TRY', '100,00', '20',
                'TRY', '150,00', '20',
                'TED-KOD-001', '50', '5', 'Açıklama',
            ],
        },
        'resim': {
            'filename': 'sablon_resim_import.xlsx',
            'headers': ['Stok Kodu', 'Resim URL'],
            'sample': ['STK-001', 'https://example.com/urun-resmi.jpg'],
        },
        'kasa': {
            'filename': 'sablon_kasa_hareket.xlsx',
            'headers': ['Tarih (GG.AA.YYYY)', 'Hareket Tipi (tahsilat/odeme/giris/cikis)',
                        'Tutar', 'Açıklama', 'Belge No', 'Cari Ünvan'],
            'sample': ['01.01.2026', 'tahsilat', '1500,00', 'Müşteri tahsilatı', 'MK-001', 'Örnek Müşteri'],
        },
    }

    if tip not in sablonlar:
        flash('Geçersiz şablon tipi.', 'danger')
        return redirect(url_for('import_export.index'))

    s = sablonlar[tip]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tip.capitalize()

    header_fill = PatternFill('solid', fgColor='1e2a3a')
    header_font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(s['headers'], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 18)

    for col, v in enumerate(s['sample'], 1):
        ws.cell(row=2, column=col, value=v)

    # Stok şablonuna geçerli değerler satırı ekle
    if tip == 'stok':
        aciklama_satiri = {
            'Ürün Tipi': 'stok veya hizmet',
            'Alış Para Birimi': 'TRY / USD / EUR',
            'Satış Para Birimi': 'TRY / USD / EUR',
            'Alış KDV Oranı': '0 / 1 / 10 / 20',
            'Satış KDV Oranı': '0 / 1 / 10 / 20',
        }
        note_font = Font(italic=True, color='888888')
        for col, h in enumerate(s['headers'], 1):
            if h in aciklama_satiri:
                cell = ws.cell(row=3, column=col, value=aciklama_satiri[h])
                cell.font = note_font

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=s['filename'],
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── Cari import ─────────────────────────────────────────────────────────────

@bp.route('/cari', methods=['POST'])
@login_required
def import_cari():
    f = request.files.get('dosya')
    if not f or not _ext_ok(f.filename):
        flash('Geçerli bir Excel (.xlsx) veya CSV dosyası seçin.', 'danger')
        return redirect(url_for('import_export.index'))

    guncelle = request.form.get('guncelle') == '1'
    rows = _read_workbook(f)
    eklenen = guncellenen = atlanan = 0

    for row in rows:
        unvan = _find(row, 'ünvan', 'unvan', 'ad', 'firma adı', 'cari adi', 'müşteri adı')
        if not unvan:
            atlanan += 1
            continue
        unvan = str(unvan).strip()

        kod = _find(row, 'cari kodu', 'kod', 'cari kod')
        tip_raw = str(_find(row, 'cari tipi (musteri/tedarikci/her_ikisi)', 'cari tipi', 'tip', 'tür') or 'musteri').lower()
        tip_map = {'müşteri': 'musteri', 'tedarikçi': 'tedarikci', 'her ikisi': 'her_ikisi',
                   'musteri': 'musteri', 'tedarikci': 'tedarikci', 'her_ikisi': 'her_ikisi'}
        tip = tip_map.get(tip_raw, 'musteri')

        mevcut = None
        if kod:
            mevcut = Cari.query.filter_by(cari_kodu=str(kod).strip()).first()
        if not mevcut and guncelle:
            mevcut = Cari.query.filter(Cari.unvan.ilike(unvan)).first()

        if mevcut and not guncelle:
            atlanan += 1
            continue

        if not mevcut:
            mevcut = Cari(cari_kodu=str(kod).strip() if kod else next_sequence('CRR', Cari, 'cari_kodu'))
            db.session.add(mevcut)
            eklenen += 1
        else:
            guncellenen += 1

        mevcut.unvan = unvan
        mevcut.cari_tipi = tip
        mevcut.vergi_no = str(_find(row, 'vergi no', 'vkn') or '').strip() or None
        mevcut.tckn = str(_find(row, 'tckn', 'tc kimlik no') or '').strip() or None
        mevcut.vergi_dairesi = str(_find(row, 'vergi dairesi') or '').strip() or None
        mevcut.telefon = str(_find(row, 'telefon', 'tel') or '').strip() or None
        mevcut.email = str(_find(row, 'e-posta', 'email', 'eposta', 'mail') or '').strip() or None
        mevcut.adres = str(_find(row, 'adres') or '').strip() or None
        mevcut.sehir = str(_find(row, 'şehir', 'sehir', 'il') or '').strip() or None
        mevcut.acik_hesap_limiti = _to_float(_find(row, 'açık hesap limiti', 'acik hesap limiti'))
        mevcut.odeme_vadesi = int(_to_float(_find(row, 'ödeme vadesi (gün)', 'vade'), 0))

    db.session.commit()
    flash(f'Cari import: {eklenen} eklendi, {guncellenen} güncellendi, {atlanan} atlandı.', 'success')
    return redirect(url_for('import_export.index'))


# ─── Stok import ─────────────────────────────────────────────────────────────

@bp.route('/stok', methods=['POST'])
@login_required
def import_stok():
    f = request.files.get('dosya')
    if not f or not _ext_ok(f.filename):
        flash('Geçerli bir Excel (.xlsx) veya CSV dosyası seçin.', 'danger')
        return redirect(url_for('import_export.index'))

    guncelle = request.form.get('guncelle') == '1'
    acilis_hareket = request.form.get('acilis_hareket') == '1'
    varsayilan_pb = request.form.get('varsayilan_pb', 'TRY').upper()
    if varsayilan_pb not in ('TRY', 'USD', 'EUR'):
        varsayilan_pb = 'TRY'
    rows = _read_workbook(f)
    eklenen = guncellenen = atlanan = 0

    kat_cache = {}

    for row in rows:
        ad = _find(row, 'ürün adı', 'urun adi', 'ad', 'stok adı', 'malzeme adı')
        if not ad:
            atlanan += 1
            continue
        ad = str(ad).strip()

        kod = _find(row, 'stok kodu', 'kod', 'stok kod')
        barkod = str(_find(row, 'barkod', 'barkod no') or '').strip() or None

        mevcut = None
        if kod:
            mevcut = Stok.query.filter_by(stok_kodu=str(kod).strip()).first()
        if not mevcut and barkod:
            mevcut = Stok.query.filter_by(barkod=barkod).first()
        if not mevcut and guncelle:
            mevcut = Stok.query.filter(Stok.ad.ilike(ad)).first()

        if mevcut and not guncelle:
            atlanan += 1
            continue

        kat_ad = str(_find(row, 'kategori', 'grup') or '').strip()
        kat_id = None
        if kat_ad:
            if kat_ad not in kat_cache:
                kat = StokKategori.query.filter(StokKategori.ad.ilike(kat_ad)).first()
                if not kat:
                    kat = StokKategori(ad=kat_ad)
                    db.session.add(kat)
                    db.session.flush()
                kat_cache[kat_ad] = kat.id
            kat_id = kat_cache[kat_ad]

        is_yeni = False
        if not mevcut:
            mevcut = Stok(stok_kodu=str(kod).strip() if kod else next_sequence('STK', Stok, 'stok_kodu'))
            db.session.add(mevcut)
            eklenen += 1
            is_yeni = True
        else:
            guncellenen += 1

        mevcut.ad = ad
        mevcut.barkod = barkod
        mevcut.kategori_id = kat_id
        mevcut.birim = str(_find(row, 'birim', 'ölçü birimi') or 'Adet').strip()
        mevcut.urun_tipi = str(_find(row, 'ürün tipi', 'urun tipi', 'tip') or 'stok').strip().lower()
        if mevcut.urun_tipi not in ('stok', 'hizmet'):
            mevcut.urun_tipi = 'stok'
        mevcut.marka = str(_find(row, 'marka', 'brand') or '').strip() or None
        mevcut.tedarikci_kodu = str(_find(row, 'tedarikçi kodu', 'tedarikci kodu') or '').strip() or None
        mevcut.alis_fiyati = _to_float(_find(row, 'alış fiyatı', 'alis fiyati', 'maliyet'))
        mevcut.satis_fiyati = _to_float(_find(row, 'satış fiyatı', 'satis fiyati', 'liste fiyatı'))
        mevcut.alis_kdv_orani = _to_float(_find(row, 'alış kdv oranı (%)', 'alis kdv', 'kdv oranı (%)'), 20)
        mevcut.kdv_orani = _to_float(_find(row, 'satış kdv oranı (%)', 'satis kdv', 'kdv oranı (%)'), 20)
        _genel_pb = str(_find(row, 'para birimi', 'currency', 'pb') or '').strip().upper()
        alis_pb = str(_find(row, 'alış para birimi', 'alis para birimi', 'alis pb') or _genel_pb or varsayilan_pb).strip().upper()
        mevcut.alis_para_birimi = alis_pb if alis_pb in ('TRY', 'USD', 'EUR') else varsayilan_pb
        satis_pb = str(_find(row, 'satış para birimi', 'satis para birimi', 'satis pb') or _genel_pb or varsayilan_pb).strip().upper()
        mevcut.satis_para_birimi = satis_pb if satis_pb in ('TRY', 'USD', 'EUR') else varsayilan_pb
        mevcut.min_stok = _to_float(_find(row, 'min stok', 'minimum stok'), 0)

        stok_miktari = _to_float(_find(row, 'mevcut stok', 'stok miktarı', 'stok miktar', 'miktar'), 0)
        if is_yeni and acilis_hareket and stok_miktari > 0:
            db.session.flush()
            h = StokHareket(
                stok_id=mevcut.id,
                hareket_tipi='giris',
                miktar=stok_miktari,
                birim_fiyat=mevcut.alis_fiyati,
                toplam_tutar=float(mevcut.alis_fiyati or 0) * stok_miktari,
                aciklama='Import açılış stoku',
                user_id=current_user.id,
            )
            mevcut.stok_miktari = stok_miktari
            db.session.add(h)

    db.session.commit()
    flash(f'Stok import: {eklenen} eklendi, {guncellenen} güncellendi, {atlanan} atlandı.', 'success')
    return redirect(url_for('import_export.index'))


# ─── Kasa hareketi import ─────────────────────────────────────────────────────

@bp.route('/kasa', methods=['POST'])
@login_required
def import_kasa():
    f = request.files.get('dosya')
    if not f or not _ext_ok(f.filename):
        flash('Geçerli bir Excel (.xlsx) veya CSV dosyası seçin.', 'danger')
        return redirect(url_for('import_export.index'))

    kasa_id = request.form.get('kasa_id', type=int)
    kasa = KasaHesap.query.get_or_404(kasa_id)
    rows = _read_workbook(f)
    eklenen = atlanan = 0

    tip_map = {
        'tahsilat': 'tahsilat', 'tahsılat': 'tahsilat',
        'ödeme': 'odeme', 'odeme': 'odeme', 'ödeme (çıkış)': 'odeme',
        'giriş': 'giris', 'giris': 'giris',
        'çıkış': 'cikis', 'cikis': 'cikis',
        'virman': 'giris',
    }

    for row in rows:
        tutar_raw = _find(row, 'tutar', 'miktar', 'borç', 'alacak')
        if tutar_raw is None:
            atlanan += 1
            continue
        tutar = _to_float(tutar_raw)
        if tutar <= 0:
            atlanan += 1
            continue

        tip_raw = str(_find(row, 'hareket tipi (tahsilat/odeme/giris/cikis)', 'hareket tipi', 'tip', 'tür') or 'giris').lower().strip()
        tip = tip_map.get(tip_raw, 'giris')
        tarih = _to_date(_find(row, 'tarih (gg.aa.yyyy)', 'tarih', 'işlem tarihi'))
        aciklama = str(_find(row, 'açıklama', 'aciklama', 'açıklama/detay') or '').strip()
        belge_no = str(_find(row, 'belge no', 'belge numarası', 'fiş no') or '').strip() or None

        h = KasaHareket(
            kasa_id=kasa.id,
            hareket_tipi=tip,
            tutar=tutar,
            aciklama=aciklama,
            tarih=tarih,
            belge_no=belge_no,
            user_id=current_user.id,
        )
        if tip in ('tahsilat', 'giris'):
            kasa.bakiye = float(kasa.bakiye or 0) + tutar
        else:
            kasa.bakiye = float(kasa.bakiye or 0) - tutar
        db.session.add(h)
        eklenen += 1

    db.session.commit()
    flash(f'Kasa hareketi import: {eklenen} kayıt eklendi, {atlanan} atlandı.', 'success')
    return redirect(url_for('import_export.index'))


# ─── Toplu resim import ───────────────────────────────────────────────────────

def _resim_url_indir(url, klasor):
    """Downloads image from URL into klasor. Returns filename or None."""
    import urllib.request
    import urllib.parse
    if not url or not str(url).startswith(('http://', 'https://')):
        return None
    url = str(url).strip()
    try:
        parsed_path = urllib.parse.urlparse(url).path
        ext = os.path.splitext(parsed_path)[1].lower().split('?')[0]
        izinli = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}
        if ext not in izinli:
            ext = '.jpg'
        dosya_adi = f"{uuid.uuid4().hex}{ext}"
        tam_yol = os.path.join(klasor, dosya_adi)
        headers = {'User-Agent': 'Mozilla/5.0'}
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=15) as resp:
            ct = resp.headers.get('Content-Type', '').split(';')[0].strip()
            ct_ext = {'image/jpeg': '.jpg', 'image/png': '.png',
                      'image/gif': '.gif', 'image/webp': '.webp'}.get(ct)
            if ct_ext and ct_ext in izinli:
                dosya_adi = f"{uuid.uuid4().hex}{ct_ext}"
                tam_yol = os.path.join(klasor, dosya_adi)
            with open(tam_yol, 'wb') as f:
                f.write(resp.read(5 * 1024 * 1024))
        return dosya_adi
    except Exception:
        return None


@bp.route('/resim', methods=['POST'])
@login_required
def import_resim():
    f = request.files.get('dosya')
    if not f or not _ext_ok(f.filename):
        flash('Geçerli bir Excel (.xlsx) veya CSV dosyası seçin.', 'danger')
        return redirect(url_for('import_export.index'))

    rows = _read_workbook(f)
    klasor = os.path.join(current_app.root_path, 'static', 'uploads', 'stok')
    os.makedirs(klasor, exist_ok=True)

    guncellenen = atlanan = hata = 0

    for row in rows:
        kod = _find(row, 'stok kodu', 'kod', 'stok kod', 'sku', 'barcode', 'barkod')
        url = _find(row, 'resim url', 'resim', 'image url', 'image', 'url', 'foto', 'fotograf')
        if not kod or not url:
            atlanan += 1
            continue
        kod = str(kod).strip()
        url = str(url).strip()

        stok = Stok.query.filter_by(stok_kodu=kod).first()
        if not stok:
            # Barkod ile de dene
            stok = Stok.query.filter_by(barkod=kod).first()
        if not stok:
            atlanan += 1
            continue

        dosya_adi = _resim_url_indir(url, klasor)
        if not dosya_adi:
            hata += 1
            continue

        # Eski resmi sil
        if stok.resim:
            eski = os.path.join(current_app.root_path, 'static', stok.resim)
            if os.path.exists(eski):
                os.remove(eski)

        stok.resim = f"uploads/stok/{dosya_adi}"
        guncellenen += 1

    db.session.commit()

    mesaj = f'Resim import: {guncellenen} ürün güncellendi'
    if atlanan:
        mesaj += f', {atlanan} atlandı (kod bulunamadı)'
    if hata:
        mesaj += f', {hata} URL indirilemedi'
    flash(mesaj, 'success' if guncellenen else 'warning')
    return redirect(url_for('import_export.index'))
