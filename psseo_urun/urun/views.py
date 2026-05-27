import io
import openpyxl
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Q

from .models import Urun, Kategori
from .forms import UrunForm, UrunImportForm


def urun_liste(request):
    q     = request.GET.get('q', '').strip()
    marka = request.GET.get('marka', '').strip()
    qs    = Urun.objects.select_related('kategori').all()
    if q:
        qs = qs.filter(Q(urun_kodu__icontains=q) | Q(ad__icontains=q) | Q(marka__icontains=q))
    if marka:
        qs = qs.filter(marka__icontains=marka)
    markalar = Urun.objects.values_list('marka', flat=True).distinct().order_by('marka')
    return render(request, 'urun/liste.html', {
        'urunler': qs, 'q': q, 'marka': marka, 'markalar': markalar
    })


def urun_detay(request, pk):
    urun      = get_object_or_404(Urun, pk=pk)
    icerikleri = urun.seo_icerikleri.all()
    # Her tip için en son onaylı veya en son üretilmiş içerik
    tip_icerik = {}
    for tip, _ in [('urun_adi',''), ('ozet',''), ('detay',''), ('meta_baslik',''),
                   ('meta_aciklama',''), ('anahtar_kelime','')]:
        onaylandi = icerikleri.filter(tip=tip, onaylandi=True).first()
        son       = icerikleri.filter(tip=tip).first()
        tip_icerik[tip] = {'onaylandi': onaylandi, 'son': son,
                           'gecmis': icerikleri.filter(tip=tip)[:5]}
    tipler = [('urun_adi','Ürün Adı Önerisi'),('ozet','Özet'),('detay','Detay Açıklama'),
              ('meta_baslik','Meta Başlık'),('meta_aciklama','Meta Açıklama'),
              ('anahtar_kelime','Anahtar Kelimeler')]
    return render(request, 'urun/detay.html', {'urun': urun, 'tip_icerik': tip_icerik, 'tipler': tipler})


def urun_ekle(request):
    form = UrunForm(request.POST or None)
    if form.is_valid():
        urun = form.save()
        messages.success(request, f'Ürün eklendi: {urun.urun_kodu}')
        return redirect('urun:detay', pk=urun.pk)
    return render(request, 'urun/form.html', {'form': form, 'baslik': 'Ürün Ekle'})


def urun_duzenle(request, pk):
    urun = get_object_or_404(Urun, pk=pk)
    form = UrunForm(request.POST or None, instance=urun)
    if form.is_valid():
        form.save()
        messages.success(request, 'Ürün güncellendi.')
        return redirect('urun:detay', pk=urun.pk)
    return render(request, 'urun/form.html', {'form': form, 'baslik': 'Ürün Düzenle', 'urun': urun})


def urun_sil(request, pk):
    urun = get_object_or_404(Urun, pk=pk)
    if request.method == 'POST':
        urun.delete()
        messages.success(request, 'Ürün silindi.')
        return redirect('urun:liste')
    return render(request, 'urun/sil_onayla.html', {'urun': urun})


def urun_import(request):
    form = UrunImportForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        dosya   = request.FILES['dosya']
        wb      = openpyxl.load_workbook(dosya)
        ws      = wb.active
        eklendi = 0
        hatalar = []
        basliklar = [str(c.value or '').strip().lower() for c in ws[1]]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            satir = dict(zip(basliklar, row))
            kodu  = str(satir.get('urun_kodu') or satir.get('ürün kodu') or '').strip()
            if not kodu:
                hatalar.append(f'Satır {i}: ürün_kodu boş, atlandı')
                continue
            urun, yeni = Urun.objects.get_or_create(urun_kodu=kodu)
            urun.ad             = str(satir.get('ad') or satir.get('ürün adı') or urun.ad or '')
            urun.marka          = str(satir.get('marka') or urun.marka or '')
            urun.ham_aciklama   = str(satir.get('ham_aciklama') or satir.get('açıklama') or urun.ham_aciklama or '')
            urun.teknik_ozellik = str(satir.get('teknik_ozellik') or satir.get('teknik özellikler') or urun.teknik_ozellik or '')
            urun.kaynak_url     = str(satir.get('kaynak_url') or satir.get('url') or urun.kaynak_url or '')
            urun.save()
            eklendi += 1
        mesaj = f'{eklendi} ürün içe aktarıldı.'
        if hatalar:
            mesaj += f' {len(hatalar)} satır atlandı.'
        messages.success(request, mesaj)
        return redirect('urun:liste')
    return render(request, 'urun/import.html', {'form': form})


def urun_export(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ürünler'
    ws.append(['urun_kodu', 'marka', 'ad', 'kategori', 'barkod', 'fiyat',
               'ham_aciklama', 'teknik_ozellik', 'kaynak_url'])
    for u in Urun.objects.select_related('kategori').all():
        ws.append([u.urun_kodu, u.marka, u.ad,
                   u.kategori.ad if u.kategori else '',
                   u.barkod, u.fiyat,
                   u.ham_aciklama, u.teknik_ozellik, u.kaynak_url])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="urunler.xlsx"'
    return resp


def urun_scrape(request):
    """AJAX veya form POST — URL'den ürün bilgisi çeker, JSON döner."""
    import json
    from seo.services.scraper_service import scrape_url
    url = request.POST.get('url', '').strip()
    if not url:
        return HttpResponse(json.dumps({'hata': 'URL boş'}), content_type='application/json')
    try:
        veri = scrape_url(url)
        return HttpResponse(json.dumps(veri, ensure_ascii=False), content_type='application/json')
    except Exception as e:
        return HttpResponse(json.dumps({'hata': str(e)}), content_type='application/json')
